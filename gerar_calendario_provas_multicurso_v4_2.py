from __future__ import annotations

from base64 import b64decode
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
import unicodedata

import pandas as pd
from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURAÇÃO
# ============================================================

VERSAO_ALGORITMO = "4.2"

# Novidade da versão 4.2:
# o layout oficial é criado pelo próprio script com uma coluna para
# cada data de prova. Dias da semana repetidos não são consolidados.

ORDEM_DIAS = {
    "Segunda": 1,
    "Terça": 2,
    "Quarta": 3,
    "Quinta": 4,
    "Sexta": 5,
    "Sábado": 6,
    "Domingo": 7,
}

DIA_PYTHON_PARA_PT = {
    0: "Segunda",
    1: "Terça",
    2: "Quarta",
    3: "Quinta",
    4: "Sexta",
    5: "Sábado",
    6: "Domingo",
}

DIAS_VISUAIS_PADRAO = [
    "Segunda",
    "Terça",
    "Quarta",
    "Quinta",
    "Sexta",
]

#: Marcador usado quando a grade não informa a sala da prova.
SALA_INDEFINIDA = "xx"

# Salas que não identificam uma turma física: duas ofertas que só
# coincidem por um destes rótulos não podem ser tratadas como a
# mesma aula.
SALAS_SEM_IDENTIDADE: frozenset[str] = frozenset(
    {
        "",
        "x",
        "xx",
        "-",
        "teams",
        "online",
        "a definir",
        "a combinar",
    }
)

ORDEM_HORARIOS = [
    "07:30",
    "09:50",
    "13:30",
    "15:50",
    "18:40",
    "20:40",
]

# Penalidades:
# 2 provas no mesmo dia é ruim;
# 3 provas é muito pior;
# 4 ou mais recebe penalidade fortíssima.
# Quando um dia de aula não tem nenhuma data no calendário de provas
# (na AS, por exemplo, não há sexta-feira), as provas desse dia são
# realocadas para um dos dias abaixo, mantendo o mesmo horário. O
# solver escolhe qual, pelo mesmo critério de conflito das demais.
DIAS_SUBSTITUTOS_PADRAO: dict[str, tuple[str, ...]] = {
    "Sexta": (
        "Terça",
        "Quarta",
        "Quinta",
    ),
}

# A prova realocada não fica no horário da aula: ela vai para o
# último horário do dia. Os horários de origem já estão ocupados
# pelas disciplinas que têm aula no dia de destino, e o professor da
# disciplina realocada costuma ser o mesmo — ele não pode estar em
# duas salas ao mesmo tempo. Use None para manter o horário da aula.
# Prova de disciplina noturna realocada para outro dia. O 18:40 do
# dia de destino já está ocupado pelas disciplinas que têm aula
# nele, então a que chega de fora fica com o segundo tempo.
HORA_DIA_SUBSTITUTO: str | None = "20:40"

# Prova de disciplina noturna no seu próprio dia de aula: sempre no
# primeiro tempo da noite, independentemente de a aula ser 18:40 ou
# 20:40.
HORA_NOTURNA_PADRAO: str | None = "18:40"

# As duas regras acima valem só para a noite. Uma disciplina diurna
# mantém o horário da aula em qualquer caso: a turma dela não tem
# como fazer prova à noite.
HORARIOS_NOTURNOS: frozenset[str] = frozenset(
    {
        "18:40",
        "20:40",
    }
)

PESO_SEGUNDA_PROVA = 100
PESO_TERCEIRA_PROVA = 10_000
PESO_QUARTA_OU_MAIS = 100_000

# Duas provas do mesmo grupo no mesmo dia E no mesmo horário são
# fisicamente impossíveis para o aluno, não apenas indesejáveis.
# O peso domina todos os demais termos do objetivo, de modo que o
# solver só aceita uma sobreposição quando não existe alternativa.
PESO_MESMO_HORARIO = 1_000_000_000

TEMPO_MAXIMO_SOLVER_SEGUNDOS = 60
NUM_SEARCH_WORKERS = 8

# Desde a versão 4.1, a concentração por professor é um critério
# lexicograficamente secundário: ela nunca pode piorar a distribuição
# de provas dos alunos. Entre calendários igualmente bons para os
# alunos, o solver minimiza primeiro os dias adicionais usados por cada
# professor e, depois, o intervalo entre a primeira e a última prova.


# ============================================================
# MODELOS
# ============================================================

@dataclass(frozen=True)
class OpcaoSemanal:
    dia: str
    hora: str


@dataclass(frozen=True)
class OpcaoProva:
    dia: str
    hora: str
    data_prova: date | None = None


@dataclass
class OfertaAcademica:
    """
    Uma oferta é identificada operacionalmente por:
        codigo + turma + agenda + professor normalizado

    codigo + turma continua sendo a identidade acadêmica principal.
    A agenda e o professor desambiguam ofertas independentes. O nome
    da disciplina NÃO participa da identidade: o mesmo código pode ter
    nomes diferentes em cursos distintos e continuar sendo uma única
    oferta quando turma, agenda e professor coincidem.

    Uma única oferta pode atender vários grupos de alunos:
        (curso, periodo, cod_turma)

    O grupo é a turma real de alunos. Duas turmas do mesmo período
    (AA e AB) são grupos DIFERENTES: provas no mesmo dia para turmas
    diferentes não são choque e não devem ser penalizadas.

    Quando a grade não informa cod_turma, o grupo cai em
    (curso, periodo, "") — comportamento conservador idêntico ao
    anterior, em que todo o período é tratado como um único grupo.

    Exemplo:
        IBM3050 + turma 8001
        grupos = {
            ("CDIA", 3, "AA"),
            ("ECOMP", 3, "AA"),
            ("ESW", 4, "AB"),
        }

    O solver escolhe UMA única prova para essa oferta, e a data
    escolhida conta como prova para todos os grupos atendidos.
    """
    id: int
    codigo: str
    materia: str
    turma: str
    professores: set[str] = field(default_factory=set)
    salas: set[str] = field(default_factory=set)
    grupos: set[tuple[str, int, str]] = field(default_factory=set)
    materias_por_grupo: dict[
        tuple[str, int, str],
        str,
    ] = field(default_factory=dict)
    opcoes_semanais: set[OpcaoSemanal] = field(default_factory=set)


@dataclass(frozen=True)
class MetricasOtimizacao:
    objetivo_alunos: int
    dias_adicionais_professores: int
    intervalo_total_professores: int
    professores_considerados: int
    concentracao_por_professor_ativa: bool
    concentracao_por_professor_otimizada: bool


# ============================================================
# NORMALIZAÇÃO
# ============================================================

def limpar_texto(valor: Any) -> str:
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def limpar_identificador(valor: Any) -> str:
    if pd.isna(valor):
        return ""

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    texto = str(valor).strip()

    try:
        numero = float(texto)
        if numero.is_integer():
            return str(int(numero))
    except ValueError:
        pass

    return texto


def normalizar_codigo(valor: Any) -> str:
    codigo = limpar_texto(valor).upper()

    if not codigo:
        return ""

    # Remove espaços internos acidentais: "IBM 3050" -> "IBM3050".
    return "".join(codigo.split())


def normalizar_hora(valor: Any) -> str:
    if pd.isna(valor):
        return ""

    if hasattr(valor, "hour") and hasattr(valor, "minute"):
        return f"{int(valor.hour):02d}:{int(valor.minute):02d}"

    texto = str(valor).strip().replace(",", ".")

    if not texto:
        return ""

    if ":" in texto:
        partes = texto.split(":")
        hora = int(partes[0])
        minuto = int(partes[1])

    elif "." in texto:
        hora_txt, minuto_txt = texto.split(".", maxsplit=1)
        hora = int(hora_txt)

        # 7.30 lido como float pode virar 7.3.
        if len(minuto_txt) == 1:
            minuto_txt += "0"

        minuto = int(minuto_txt[:2])

    else:
        hora = int(texto)
        minuto = 0

    if not 0 <= hora <= 23:
        raise ValueError(f"Hora inválida: {valor}")

    if not 0 <= minuto <= 59:
        raise ValueError(f"Minuto inválido: {valor}")

    return f"{hora:02d}:{minuto:02d}"


def normalizar_dia(valor: Any) -> str:
    dia = limpar_texto(valor)

    if not dia:
        return ""

    aliases = {
        "segunda": "Segunda",
        "segunda-feira": "Segunda",
        "terca": "Terça",
        "terça": "Terça",
        "terca-feira": "Terça",
        "terça-feira": "Terça",
        "quarta": "Quarta",
        "quarta-feira": "Quarta",
        "quinta": "Quinta",
        "quinta-feira": "Quinta",
        "sexta": "Sexta",
        "sexta-feira": "Sexta",
        "sabado": "Sábado",
        "sábado": "Sábado",
        "domingo": "Domingo",
    }

    chave = dia.casefold()

    if chave in aliases:
        return aliases[chave]

    if dia in ORDEM_DIAS:
        return dia

    raise ValueError(f"Dia inválido: {dia}")


def normalizar_data(valor: str | date | datetime) -> date:
    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()

    for formato in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(
                texto,
                formato,
            ).date()
        except ValueError:
            pass

    raise ValueError(
        f"Data inválida: {valor!r}. "
        "Use YYYY-MM-DD ou DD/MM/YYYY."
    )


def normalizar_datas_provas(
    datas_provas: Iterable[str | date | datetime] | None,
) -> list[date] | None:
    if datas_provas is None:
        return None

    datas = sorted(
        {
            normalizar_data(valor)
            for valor in datas_provas
        }
    )

    if not datas:
        raise ValueError(
            "datas_provas foi informado, mas a lista está vazia."
        )

    return datas


def ordenar_datas_por_dia_semana(
    datas_provas: list[date],
) -> list[date]:
    """Gira as colunas da tabela para que ela comece na segunda.

    A ordem cronológica é preservada: a lista apenas é rotacionada
    para o ponto em que uma nova semana começa, ou seja, o primeiro
    dia que é anterior na semana ao dia que vem antes dele. As datas
    que ficaram para trás vão para o fim.

        AP1  Qui 24/09, Sex 25/09, Seg 28/09, Ter 29/09, Qua 30/09
          -> Seg 28/09, Ter 29/09, Qua 30/09, Qui 24/09, Sex 25/09

    Um calendário que já começa na segunda não é alterado.
    """
    datas = sorted(
        datas_provas
    )

    for indice in range(
        1,
        len(datas),
    ):
        dia_atual = ORDEM_DIAS[
            nome_dia(
                datas[indice]
            )
        ]
        dia_anterior = ORDEM_DIAS[
            nome_dia(
                datas[indice - 1]
            )
        ]

        if dia_atual < dia_anterior:
            return (
                datas[indice:]
                + datas[:indice]
            )

    return datas


def nome_dia(data_prova: date) -> str:
    return DIA_PYTHON_PARA_PT[
        data_prova.weekday()
    ]


def formatar_data_br(
    data_prova: date | None,
) -> str:
    if data_prova is None:
        return ""

    return data_prova.strftime(
        "%d/%m/%Y"
    )


def periodo_letivo_padrao(
    datas_provas: list[date] | None,
) -> str:
    """
    Deduz o rótulo "AAAA.S" a partir das datas de prova.

    Provas até junho pertencem ao 1º semestre; a partir de julho,
    ao 2º. Serve apenas como padrão quando periodo_letivo não é
    informado explicitamente.
    """
    referencia = (
        datas_provas[0]
        if datas_provas
        else date.today()
    )

    semestre = (
        1
        if referencia.month <= 6
        else 2
    )

    return f"{referencia.year}.{semestre}"


def normalizar_nome_comparacao(
    texto: str,
) -> str:
    texto_sem_acentos = "".join(
        caractere
        for caractere in unicodedata.normalize(
            "NFKD",
            texto,
        )
        if not unicodedata.combining(
            caractere
        )
    )

    return " ".join(
        texto_sem_acentos.casefold().split()
    )


def normalizar_cod_turma(valor: Any) -> str:
    """
    Código não numérico da turma (AA, AB, AC...).

    Coluna opcional: grades que não a possuem continuam válidas e
    caem no comportamento antigo (um grupo por curso + período).
    """
    return (
        limpar_identificador(valor)
        .upper()
    )


def rotulo_grupo(
    grupo: tuple[str, int, str],
) -> str:
    """
    Rótulo legível de um grupo de alunos.

    Sem cod_turma o rótulo é idêntico ao das versões anteriores,
    para não quebrar a leitura das planilhas antigas.
    """
    curso, periodo, cod_turma = grupo

    if cod_turma:
        return f"{curso}-{periodo}-{cod_turma}"

    return f"{curso}-{periodo}"


def identificador_grupo(
    grupo: tuple[str, int, str],
) -> str:
    """
    Versão do rótulo segura para nomes de variáveis do CP-SAT.
    """
    return (
        rotulo_grupo(grupo)
        .replace(" ", "_")
        .replace("-", "_")
    )


# ============================================================
# LEITURA
# ============================================================

def carregar_grade(
    caminho: str | Path,
) -> pd.DataFrame:
    """
    O CSV/XLSX deve possuir EXATAMENTE a informação necessária
    para identificar a oferta e os grupos atendidos.

    Colunas esperadas:
        codigo
        materia
        turma
        dia_1
        hora_1
        dia_2
        hora_2
        periodo
        professor
        curso

    Colunas opcionais:
        cod_turma   código não numérico da turma (AA, AB, AC...)
        sala        sala da prova; ausente ou vazia vira "xx"

    A sala não participa da identidade da oferta nem das restrições
    do solver: é apenas transportada até o calendário oficial. Assim
    a coordenação pode preencher as salas direto no CSV e regerar,
    sem mexer em nenhum outro arquivo.

    O cod_turma identifica a TURMA REAL de alunos e é o que separa
    os grupos dentro de um mesmo período. Se a coluna não existir
    (ou vier vazia), o arquivo continua válido e todo o período é
    tratado como um único grupo, como nas versões anteriores.

    Atenção: a turma numérica NÃO serve como substituta. Uma mesma
    turma de alunos pode cursar disciplinas sob números diferentes
    quando a oferta é compartilhada com outro curso.
    """
    caminho = Path(caminho)

    if not caminho.exists():
        raise FileNotFoundError(
            f"Arquivo não encontrado: {caminho}"
        )

    if caminho.suffix.lower() == ".csv":
        df = pd.read_csv(caminho)

    elif caminho.suffix.lower() in {
        ".xlsx",
        ".xls",
    }:
        excel = pd.ExcelFile(caminho)

        sheet_name = (
            "Grade"
            if "Grade" in excel.sheet_names
            else 0
        )

        df = pd.read_excel(
            caminho,
            sheet_name=sheet_name,
        )

    else:
        raise ValueError(
            "Formato não suportado. Use CSV, XLSX ou XLS."
        )

    df.columns = [
        str(coluna)
        .strip()
        .lower()
        for coluna in df.columns
    ]

    colunas_obrigatorias = {
        "codigo",
        "materia",
        "turma",
        "dia_1",
        "hora_1",
        "dia_2",
        "hora_2",
        "periodo",
        "professor",
        "curso",
    }

    faltantes = (
        colunas_obrigatorias
        - set(df.columns)
    )

    if faltantes:
        raise ValueError(
            "Colunas obrigatórias ausentes: "
            + ", ".join(
                sorted(faltantes)
            )
        )

    df = df.copy()

    if "cod_turma" not in df.columns:
        df["cod_turma"] = ""

    df["cod_turma"] = (
        df["cod_turma"]
        .apply(normalizar_cod_turma)
    )

    if "sala" not in df.columns:
        df["sala"] = ""

    df["sala"] = (
        df["sala"]
        .apply(limpar_identificador)
        .replace("", SALA_INDEFINIDA)
    )

    df["codigo"] = (
        df["codigo"]
        .apply(normalizar_codigo)
    )

    df["materia"] = (
        df["materia"]
        .apply(limpar_texto)
    )

    df["turma"] = (
        df["turma"]
        .apply(limpar_identificador)
    )

    df["professor"] = (
        df["professor"]
        .apply(limpar_texto)
    )

    df["curso"] = (
        df["curso"]
        .apply(limpar_texto)
    )

    df["dia_1"] = (
        df["dia_1"]
        .apply(normalizar_dia)
    )

    df["dia_2"] = (
        df["dia_2"]
        .apply(normalizar_dia)
    )

    df["hora_1"] = (
        df["hora_1"]
        .apply(normalizar_hora)
    )

    df["hora_2"] = (
        df["hora_2"]
        .apply(normalizar_hora)
    )

    periodo = pd.to_numeric(
        df["periodo"],
        errors="coerce",
    )

    if periodo.isna().any():
        linhas = [
            str(indice + 2)
            for indice
            in periodo[
                periodo.isna()
            ].index
        ]

        raise ValueError(
            "Período inválido ou vazio nas linhas: "
            + ", ".join(linhas)
        )

    df["periodo"] = (
        periodo.astype(int)
    )

    erros = []

    for indice, row in df.iterrows():
        linha = indice + 2

        if not row["codigo"]:
            erros.append(
                f"Linha {linha}: código vazio."
            )

        if not row["materia"]:
            erros.append(
                f"Linha {linha}: matéria vazia."
            )

        if not row["turma"]:
            erros.append(
                f"Linha {linha}: turma vazia."
            )

        if not row["curso"]:
            erros.append(
                f"Linha {linha}: curso vazio."
            )

        if not row["dia_1"]:
            erros.append(
                f"Linha {linha}: dia_1 vazio."
            )

        if row["dia_1"] and not row["hora_1"]:
            erros.append(
                f"Linha {linha}: dia_1 preenchido sem hora_1."
            )

        if row["dia_2"] and not row["hora_2"]:
            erros.append(
                f"Linha {linha}: dia_2 preenchido sem hora_2."
            )

        if row["hora_2"] and not row["dia_2"]:
            erros.append(
                f"Linha {linha}: hora_2 preenchida sem dia_2."
            )

    if erros:
        raise ValueError(
            "Foram encontrados erros críticos na grade:\n- "
            + "\n- ".join(erros)
        )

    return df


# ============================================================
# VALIDAÇÃO DO CSV
# ============================================================

def validar_grade(
    df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Alertas gerais que não impedem a execução.
    """
    problemas = []

    for indice, row in df.iterrows():
        linha = indice + 2

        if not row["professor"]:
            problemas.append(
                {
                    "tipo": "ATENÇÃO",
                    "codigo": row["codigo"],
                    "turma": row["turma"],
                    "linha": linha,
                    "problema": "Professor vazio",
                }
            )

        if (
            row["dia_1"]
            and row["dia_2"]
            and row["dia_1"]
            == row["dia_2"]
        ):
            problemas.append(
                {
                    "tipo": "ATENÇÃO",
                    "codigo": row["codigo"],
                    "turma": row["turma"],
                    "linha": linha,
                    "problema": (
                        "dia_1 e dia_2 são iguais; conferir se "
                        "são blocos consecutivos da mesma aula."
                    ),
                }
            )

    # Preenchimento parcial de cod_turma é perigoso: as linhas sem
    # código seriam agrupadas numa turma fantasma "" e o choque
    # entre elas passaria batido.
    if df["cod_turma"].ne("").any():
        for indice in df[
            df["cod_turma"] == ""
        ].index:
            row = df.loc[indice]

            problemas.append(
                {
                    "tipo": "ATENÇÃO",
                    "codigo": row["codigo"],
                    "turma": row["turma"],
                    "linha": indice + 2,
                    "problema": (
                        "cod_turma vazio em arquivo que usa "
                        "cod_turma. Esta linha será agrupada "
                        f"como {row['curso']}-{row['periodo']} "
                        "sem turma, separada das demais."
                    ),
                }
            )

    duplicadas = (
        df.duplicated(
            keep=False
        )
    )

    for indice in df[
        duplicadas
    ].index:
        row = df.loc[indice]

        problemas.append(
            {
                "tipo": "ATENÇÃO",
                "codigo": row["codigo"],
                "turma": row["turma"],
                "linha": indice + 2,
                "problema": (
                    "Linha exatamente duplicada"
                ),
            }
        )

    if not problemas:
        return pd.DataFrame(
            columns=[
                "tipo",
                "codigo",
                "turma",
                "linha",
                "problema",
            ]
        )

    return (
        pd.DataFrame(problemas)
        .sort_values(
            [
                "codigo",
                "turma",
                "linha",
            ]
        )
        .reset_index(drop=True)
    )


def opcoes_da_linha(
    row: pd.Series,
) -> frozenset[OpcaoSemanal]:
    opcoes = set()

    if row["dia_1"]:
        opcoes.add(
            OpcaoSemanal(
                dia=row["dia_1"],
                hora=row["hora_1"],
            )
        )

    if row["dia_2"]:
        opcoes.add(
            OpcaoSemanal(
                dia=row["dia_2"],
                hora=row["hora_2"],
            )
        )

    return frozenset(opcoes)


def validar_consistencia_ofertas(
    df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Valida ocorrências com o mesmo codigo + turma.

    Regra multicurso da versão 4.1:
    - codigo + turma + agenda + professor iguais:
      representam a mesma oferta, mesmo que os nomes da disciplina
      sejam diferentes entre os cursos;
    - professor ou agenda diferentes:
      representam ofertas distintas.

    As diferenças esperadas geram INFO para auditoria, sem bloquear a
    geração do calendário.
    """
    problemas = []

    for (
        codigo,
        turma,
    ), grupo in df.groupby(
        [
            "codigo",
            "turma",
        ],
        sort=True,
    ):
        por_chave_operacional: dict[
            tuple[
                frozenset[OpcaoSemanal],
                str,
            ],
            list[pd.Series],
        ] = {}

        professores_por_agenda: dict[
            frozenset[OpcaoSemanal],
            dict[str, str],
        ] = {}

        for _, row in grupo.iterrows():
            agenda = opcoes_da_linha(
                row
            )
            professor = limpar_texto(
                row["professor"]
            )
            professor_normalizado = (
                normalizar_nome_comparacao(
                    professor
                )
            )

            por_chave_operacional.setdefault(
                (
                    agenda,
                    professor_normalizado,
                ),
                [],
            ).append(row)

            if professor_normalizado:
                professores_por_agenda.setdefault(
                    agenda,
                    {},
                ).setdefault(
                    professor_normalizado,
                    professor,
                )

        # Nomes diferentes não quebram o compartilhamento quando os
        # demais componentes da chave operacional são iguais.
        for linhas_oferta in (
            por_chave_operacional.values()
        ):
            nomes_por_chave: dict[
                str,
                str,
            ] = {}

            for row in linhas_oferta:
                materia = limpar_texto(
                    row["materia"]
                )

                if materia:
                    nomes_por_chave.setdefault(
                        normalizar_nome_comparacao(
                            materia
                        ),
                        materia,
                    )

            if len(nomes_por_chave) > 1:
                problemas.append(
                    {
                        "tipo": "INFO",
                        "codigo": codigo,
                        "turma": turma,
                        "campo": "materia",
                        "valores": " | ".join(
                            sorted(
                                nomes_por_chave.values(),
                                key=str.casefold,
                            )
                        ),
                        "problema": (
                            "Mesma oferta com nomes de disciplina "
                            "diferentes entre cursos. O sistema manterá "
                            "uma única prova e preservará o nome de cada "
                            "curso na saída."
                        ),
                    }
                )

        # Professores diferentes, ainda que a agenda seja idêntica,
        # sempre originam ofertas independentes.
        for professores_indexados in (
            professores_por_agenda.values()
        ):
            if len(professores_indexados) <= 1:
                continue

            problemas.append(
                {
                    "tipo": "INFO",
                    "codigo": codigo,
                    "turma": turma,
                    "campo": "professor",
                    "valores": " | ".join(
                        sorted(
                            professores_indexados.values(),
                            key=str.casefold,
                        )
                    ),
                    "problema": (
                        "Mesmo codigo + turma + agenda com professores "
                        "diferentes. O sistema separará automaticamente "
                        "essas ocorrências em ofertas distintas."
                    ),
                }
            )

        agendas = {
            agenda
            for agenda, _
            in por_chave_operacional
        }

        if len(agendas) > 1:
            agendas_txt = []

            for agenda in agendas:
                ordenada = sorted(
                    agenda,
                    key=lambda o: (
                        ORDEM_DIAS[o.dia],
                        o.hora,
                    ),
                )

                agendas_txt.append(
                    " / ".join(
                        f"{o.dia} {o.hora}"
                        for o in ordenada
                    )
                )

            problemas.append(
                {
                    "tipo": "INFO",
                    "codigo": codigo,
                    "turma": turma,
                    "campo": "agenda",
                    "valores": " || ".join(
                        sorted(agendas_txt)
                    ),
                    "problema": (
                        "Mesmo codigo + turma possui agendas diferentes. "
                        "O sistema separará automaticamente essas "
                        "ocorrências em ofertas distintas."
                    ),
                }
            )

    if not problemas:
        return pd.DataFrame(
            columns=[
                "tipo",
                "codigo",
                "turma",
                "campo",
                "valores",
                "problema",
            ]
        )

    return (
        pd.DataFrame(problemas)
        .sort_values(
            [
                "tipo",
                "codigo",
                "turma",
                "campo",
            ]
        )
        .reset_index(drop=True)
    )

def garantir_consistencia_critica(
    validacao_ofertas: pd.DataFrame,
) -> None:
    if validacao_ofertas.empty:
        return

    criticos = validacao_ofertas[
        validacao_ofertas[
            "tipo"
        ] == "ERRO CRÍTICO"
    ]

    if criticos.empty:
        return

    detalhes = []

    for _, row in criticos.iterrows():
        detalhes.append(
            f"- {row['codigo']} / turma {row['turma']}: "
            f"{row['problema']} {row['valores']}"
        )

    raise ValueError(
        "Há inconsistências críticas em ofertas compartilhadas.\n"
        "O calendário não será gerado enquanto houver dados que "
        "impeçam identificar com segurança codigo + turma + agenda "
        "+ professor.\n\n"
        + "\n".join(detalhes)
    )


# ============================================================
# EXCLUSÃO DE ESTUDO DIRIGIDO E ELETIVAS
# ============================================================

def separar_disciplinas_excluidas(
    df: pd.DataFrame,
    excluir: bool = True,
) -> tuple[
    pd.DataFrame,
    pd.DataFrame,
]:
    """
    (ED) = Estudo Dirigido
    (E)  = Eletiva

    Com excluir=True essas linhas ficam fora do solver. Com
    excluir=False elas são alocadas como qualquer outra disciplina
    (pedido da coordenação) e nada é separado.
    """
    if excluir:
        mascara = (
            df["materia"]
            .astype(str)
            .str.contains(
                r"\((?:ED|E)\)",
                case=False,
                regex=True,
                na=False,
            )
        )
    else:
        mascara = pd.Series(
            False,
            index=df.index,
        )

    df_excluidas = (
        df.loc[mascara]
        .copy()
        .reset_index(drop=True)
    )

    df_alocacao = (
        df.loc[~mascara]
        .copy()
        .reset_index(drop=True)
    )

    return (
        df_alocacao,
        df_excluidas,
    )


def criar_dataframe_excluidas(
    df_excluidas: pd.DataFrame,
) -> pd.DataFrame:
    """
    Consolida as excluídas pela mesma chave operacional das ofertas
    alocáveis, sem misturar professores diferentes.
    """
    colunas = [
        "codigo",
        "materia",
        "turma",
        "grupos_atendidos",
        "professores",
        "tipo",
        "motivo",
    ]

    if df_excluidas.empty:
        return pd.DataFrame(
            columns=colunas
        )

    linhas = []
    df_trabalho = df_excluidas.copy()

    df_trabalho["_professor_chave"] = (
        df_trabalho["professor"]
        .apply(
            normalizar_nome_comparacao
        )
    )

    df_trabalho["_agenda_chave"] = (
        df_trabalho.apply(
            lambda row: tuple(
                (
                    opcao.dia,
                    opcao.hora,
                )
                for opcao in sorted(
                    opcoes_da_linha(row),
                    key=lambda opcao: (
                        ORDEM_DIAS[opcao.dia],
                        opcao.hora,
                    ),
                )
            ),
            axis=1,
        )
    )

    for (
        codigo,
        turma,
        _,
        _,
    ), grupo in df_trabalho.groupby(
        [
            "codigo",
            "turma",
            "_agenda_chave",
            "_professor_chave",
        ],
        sort=False,
    ):
        materias_por_chave: dict[
            str,
            str,
        ] = {}

        for valor in grupo["materia"]:
            materia_linha = limpar_texto(
                valor
            )

            if materia_linha:
                materias_por_chave.setdefault(
                    normalizar_nome_comparacao(
                        materia_linha
                    ),
                    materia_linha,
                )

        materia = " | ".join(
            sorted(
                materias_por_chave.values(),
                key=str.casefold,
            )
        )

        professores_por_chave: dict[
            str,
            str,
        ] = {}

        for valor in grupo["professor"]:
            professor_linha = limpar_texto(
                valor
            )

            if professor_linha:
                professores_por_chave.setdefault(
                    normalizar_nome_comparacao(
                        professor_linha
                    ),
                    professor_linha,
                )

        professores = sorted(
            professores_por_chave.values(),
            key=str.casefold,
        )

        grupos = sorted(
            {
                (
                    str(row["curso"]),
                    int(row["periodo"]),
                    str(row["cod_turma"]),
                )
                for _, row
                in grupo.iterrows()
            }
        )

        materias_upper = [
            str(valor).upper()
            for valor in grupo["materia"]
        ]

        if any(
            "(ED)" in valor
            for valor in materias_upper
        ):
            tipo = "ED"
            motivo = (
                "Estudo Dirigido — não participa da "
                "alocação automática de provas."
            )
        else:
            tipo = "E"
            motivo = (
                "Eletiva — não participa da "
                "alocação automática de provas."
            )

        linhas.append(
            {
                "codigo": codigo,
                "materia": materia,
                "turma": turma,
                "grupos_atendidos": (
                    " | ".join(
                        rotulo_grupo(grupo_atendido)
                        for grupo_atendido
                        in grupos
                    )
                ),
                "professores": (
                    " | ".join(
                        professores
                    )
                ),
                "tipo": tipo,
                "motivo": motivo,
            }
        )

    return pd.DataFrame(
        linhas,
        columns=colunas,
    )


# ============================================================
# CONSOLIDAÇÃO DAS OFERTAS ACADÊMICAS
# ============================================================

def construir_ofertas(
    df: pd.DataFrame,
) -> list[OfertaAcademica]:
    """
    Consolidação multicurso.

    Identidade acadêmica principal:
        codigo + turma

    Desambiguação operacional:
        codigo + turma + agenda + professor normalizado

    Se dois cursos possuem mesmo codigo, turma, agenda e professor,
    eles compartilham uma única oferta/prova. O nome da disciplina
    pode variar entre cursos sem romper o compartilhamento.

    Se a agenda ou o professor forem diferentes, as ocorrências são
    tratadas como ofertas distintas.

    O cod_turma NÃO entra na identidade da oferta, apenas nos
    grupos atendidos. Assim, se duas turmas de alunos assistem à
    mesma aula, a prova continua sendo uma só — e o compartilhamento
    entre cursos não se rompe quando um arquivo informa cod_turma
    e outro não.
    """
    por_chave: dict[
        tuple[
            str,
            str,
            frozenset[OpcaoSemanal],
            str,
        ],
        OfertaAcademica,
    ] = {}

    for _, row in df.iterrows():
        agenda = opcoes_da_linha(row)
        professor_normalizado = (
            normalizar_nome_comparacao(
                row["professor"]
            )
        )

        chave = (
            row["codigo"],
            row["turma"],
            agenda,
            professor_normalizado,
        )

        if chave not in por_chave:
            por_chave[chave] = OfertaAcademica(
                id=len(por_chave),
                codigo=row["codigo"],
                materia=row["materia"],
                turma=row["turma"],
            )

        oferta = por_chave[chave]

        if row["professor"]:
            professor_ja_registrado = any(
                normalizar_nome_comparacao(
                    professor
                ) == professor_normalizado
                for professor in (
                    oferta.professores
                )
            )

            if not professor_ja_registrado:
                oferta.professores.add(
                    row["professor"]
                )

        if row["sala"]:
            oferta.salas.add(
                row["sala"]
            )

        grupo_alunos = (
            row["curso"],
            int(row["periodo"]),
            row["cod_turma"],
        )

        oferta.grupos.add(
            grupo_alunos
        )

        oferta.materias_por_grupo.setdefault(
            grupo_alunos,
            row["materia"],
        )

        oferta.opcoes_semanais.update(
            agenda
        )

    ofertas = list(
        por_chave.values()
    )

    for oferta in ofertas:
        if not oferta.opcoes_semanais:
            raise ValueError(
                "Oferta sem opção de prova: "
                f"{oferta.codigo} / turma {oferta.turma}"
            )

    return ofertas


def nomes_materias_oferta(
    oferta: OfertaAcademica,
) -> list[str]:
    """Nomes comerciais associados à mesma oferta entre cursos."""
    por_nome_normalizado: dict[
        str,
        str,
    ] = {}

    for materia in [
        oferta.materia,
        *oferta.materias_por_grupo.values(),
    ]:
        materia = limpar_texto(
            materia
        )

        if not materia:
            continue

        por_nome_normalizado.setdefault(
            normalizar_nome_comparacao(
                materia
            ),
            materia,
        )

    return sorted(
        por_nome_normalizado.values(),
        key=str.casefold,
    )


def texto_materias_oferta(
    oferta: OfertaAcademica,
) -> str:
    return " | ".join(
        nomes_materias_oferta(
            oferta
        )
    )


# ============================================================
# TURMAS CONJUNTAS
# ============================================================

def chave_turma_conjunta(
    oferta: OfertaAcademica,
) -> tuple[frozenset[str], frozenset[OpcaoSemanal], frozenset[str]] | None:
    """Assinatura física da aula: professor + agenda + sala.

    Devolve None quando a sala não identifica uma turma real (em
    branco, "xx", "Teams"...), porque aí a coincidência não prova
    nada.
    """
    salas = frozenset(
        limpar_texto(sala).casefold()
        for sala in oferta.salas
    )

    if (
        not salas
        or salas & SALAS_SEM_IDENTIDADE
    ):
        return None

    professores = frozenset(
        normalizar_nome_comparacao(
            professor
        )
        for professor in oferta.professores
    )

    if not professores:
        return None

    return (
        professores,
        frozenset(
            oferta.opcoes_semanais
        ),
        salas,
    )


def agrupar_turmas_conjuntas(
    ofertas: list[OfertaAcademica],
) -> list[list[OfertaAcademica]]:
    """Ofertas que, na prática, são a mesma aula.

    Mesmo professor, mesma agenda semanal e mesma sala: o professor
    não estaria em duas salas ao mesmo tempo, logo é uma turma só,
    ofertada em cursos diferentes com códigos e nomes diferentes
    (Projeto Front-End / Projeto em Ciência de Dados I, por
    exemplo). Elas fazem uma prova só, no mesmo dia e horário.
    """
    grupos: dict[Any, list[OfertaAcademica]] = {}

    for oferta in ofertas:
        chave = chave_turma_conjunta(
            oferta
        )

        if chave is None:
            continue

        grupos.setdefault(
            chave,
            [],
        ).append(oferta)

    return [
        membros
        for membros in grupos.values()
        if len(membros) > 1
    ]


def detectar_choques_de_professor(
    ofertas: list[OfertaAcademica],
) -> list[list[OfertaAcademica]]:
    """Mesmo professor e mesma agenda, mas em salas diferentes.

    Não são turmas conjuntas: é um choque na própria grade, que
    merece conferência manual. Aqui só reportamos.
    """
    grupos: dict[Any, list[OfertaAcademica]] = {}

    for oferta in ofertas:
        professores = frozenset(
            normalizar_nome_comparacao(
                professor
            )
            for professor in oferta.professores
        )

        if not professores:
            continue

        grupos.setdefault(
            (
                professores,
                frozenset(
                    oferta.opcoes_semanais
                ),
            ),
            [],
        ).append(oferta)

    choques = []

    for membros in grupos.values():
        if len(membros) < 2:
            continue

        salas = {
            frozenset(
                limpar_texto(sala).casefold()
                for sala in oferta.salas
            )
            for oferta in membros
        }

        if len(salas) > 1:
            choques.append(membros)

    return choques


# ============================================================
# OPÇÕES DE PROVA
# ============================================================

def gerar_opcoes_elegiveis(
    oferta: OfertaAcademica,
    datas_provas: list[date] | None,
    dias_substitutos: dict[str, tuple[str, ...]] | None = None,
    hora_dia_substituto: str | None = HORA_DIA_SUBSTITUTO,
    hora_noturna_padrao: str | None = HORA_NOTURNA_PADRAO,
) -> list[OpcaoProva]:
    """
    Converte os dias de aula da oferta em datas reais de prova.

    A substituição é decidida por OFERTA, não por dia de aula. Se
    pelo menos um dos dias de aula tem data no calendário, só esses
    dias valem: uma disciplina de quarta e sexta faz a prova na
    quarta. Os dias de dias_substitutos só entram quando NENHUM dia
    de aula da oferta tem data — aí a prova pode cair em qualquer um
    deles.

    O horário só é reescrito para as disciplinas NOTURNAS: no seu
    próprio dia elas fazem prova em hora_noturna_padrao (18:40),
    e quando realocadas para outro dia, em hora_dia_substituto
    (20:40), porque o 18:40 do dia de destino já é das disciplinas
    daquele dia. As diurnas mantêm o horário da aula sempre.
    """
    if datas_provas is None:
        return sorted(
            [
                OpcaoProva(
                    dia=opcao.dia,
                    hora=opcao.hora,
                    data_prova=None,
                )
                for opcao
                in oferta.opcoes_semanais
            ],
            key=lambda o: (
                ORDEM_DIAS[o.dia],
                o.hora,
            ),
        )

    substitutos = (
        dias_substitutos
        or {}
    )

    dias_com_data = {
        nome_dia(data_real)
        for data_real in datas_provas
    }

    # A oferta só recorre aos dias substitutos quando nenhum dos seus
    # dias de aula tem data no calendário.
    tem_dia_proprio = any(
        opcao.dia in dias_com_data
        for opcao
        in oferta.opcoes_semanais
    )

    opcoes = []

    for data_real in datas_provas:
        dia_real = nome_dia(
            data_real
        )

        for opcao in (
            oferta.opcoes_semanais
        ):
            noturna = (
                opcao.hora
                in HORARIOS_NOTURNOS
            )

            if tem_dia_proprio:
                elegivel = (
                    opcao.dia == dia_real
                )
                hora_alvo = (
                    hora_noturna_padrao
                )
            else:
                elegivel = (
                    dia_real
                    in substitutos.get(
                        opcao.dia,
                        (),
                    )
                )
                hora_alvo = (
                    hora_dia_substituto
                )

            hora = (
                hora_alvo
                if (
                    noturna
                    and hora_alvo
                )
                else opcao.hora
            )

            if elegivel:
                opcoes.append(
                    OpcaoProva(
                        dia=dia_real,
                        hora=hora,
                        data_prova=data_real,
                    )
                )

    return sorted(
        set(opcoes),
        key=lambda o: (
            o.data_prova,
            o.hora,
        ),
    )


# ============================================================
# OTIMIZAÇÃO
# ============================================================

def otimizar_calendario(
    ofertas: list[OfertaAcademica],
    datas_provas: list[date] | None = None,
    concentrar_por_professor: bool = True,
    dias_substitutos: dict[str, tuple[str, ...]] | None = None,
    hora_dia_substituto: str | None = HORA_DIA_SUBSTITUTO,
    hora_noturna_padrao: str | None = HORA_NOTURNA_PADRAO,
    unir_turmas_conjuntas: bool = True,
) -> tuple[
    cp_model.CpSolver,
    dict[int, OpcaoProva],
    dict[int, list[OpcaoProva]],
    MetricasOtimizacao,
]:
    """
    Cada OfertaAcademica gera UMA única variável de decisão por
    opção de prova.

    Uma oferta compartilhada aparece em vários grupos de alunos,
    mas a mesma variável é usada em todas as restrições desses
    grupos. Portanto, a prova é realmente única.
    """
    model = cp_model.CpModel()

    opcoes_por_oferta: dict[
        int,
        list[OpcaoProva],
    ] = {}

    x: dict[
        tuple[int, OpcaoProva],
        cp_model.IntVar,
    ] = {}

    ofertas_ativas: list[
        OfertaAcademica
    ] = []

    # --------------------------------------------------------
    # UMA PROVA POR OFERTA ELEGÍVEL
    # --------------------------------------------------------

    for oferta in ofertas:
        opcoes = (
            gerar_opcoes_elegiveis(
                oferta,
                datas_provas,
                dias_substitutos,
                hora_dia_substituto,
                hora_noturna_padrao,
            )
        )

        opcoes_por_oferta[
            oferta.id
        ] = opcoes

        # Sem data compatível: fica para tratamento manual.
        if not opcoes:
            continue

        ofertas_ativas.append(
            oferta
        )

        for indice, opcao in enumerate(
            opcoes
        ):
            if opcao.data_prova is None:
                id_tempo = opcao.dia
            else:
                id_tempo = (
                    opcao.data_prova
                    .isoformat()
                )

            id_tempo = (
                id_tempo
                .replace("-", "_")
            )

            hora_id = (
                opcao.hora
                .replace(":", "_")
            )

            x[
                (
                    oferta.id,
                    opcao,
                )
            ] = model.NewBoolVar(
                f"oferta_{oferta.id}_"
                f"{id_tempo}_"
                f"{hora_id}_{indice}"
            )

        model.Add(
            sum(
                x[
                    (
                        oferta.id,
                        opcao,
                    )
                ]
                for opcao in opcoes
            ) == 1
        )

    # --------------------------------------------------------
    # TURMAS CONJUNTAS: UMA AULA SÓ, UMA PROVA SÓ
    #
    # Mesmo professor, mesma agenda e mesma sala significam a mesma
    # aula, ofertada em cursos diferentes com códigos diferentes.
    # Amarrar as decisões garante que caiam sempre no mesmo dia e
    # horário — e o professor deixa de "disputar" consigo mesmo.
    # --------------------------------------------------------

    if unir_turmas_conjuntas:
        for membros in agrupar_turmas_conjuntas(
            ofertas_ativas
        ):
            referencia = membros[0]

            for outra in membros[1:]:
                for opcao in (
                    opcoes_por_oferta[
                        referencia.id
                    ]
                ):
                    chave_outra = (
                        outra.id,
                        opcao,
                    )

                    if chave_outra not in x:
                        continue

                    model.Add(
                        x[
                            (
                                referencia.id,
                                opcao,
                            )
                        ]
                        == x[chave_outra]
                    )

    # --------------------------------------------------------
    # GRUPOS: CURSO + PERÍODO + TURMA
    # --------------------------------------------------------

    grupos: dict[
        tuple[str, int, str],
        list[OfertaAcademica],
    ] = {}

    for oferta in ofertas_ativas:
        for grupo in oferta.grupos:
            grupos.setdefault(
                grupo,
                [],
            ).append(oferta)

    termos_objetivo_alunos = []

    if datas_provas is None:
        unidades_temporais: list[
            tuple[str, date | None]
        ] = [
            (
                dia,
                None,
            )
            for dia in ORDEM_DIAS
            if dia != "Domingo"
        ]
    else:
        unidades_temporais = [
            (
                nome_dia(data_real),
                data_real,
            )
            for data_real
            in datas_provas
        ]

    # --------------------------------------------------------
    # MINIMIZA CONFLITOS PARA CADA CURSO + PERÍODO + TURMA
    #
    # Cada turma de alunos é penalizada isoladamente. Duas provas
    # no mesmo dia em turmas diferentes não custam nada, porque
    # nenhum aluno faz as duas.
    # --------------------------------------------------------

    for grupo, ofertas_grupo in grupos.items():
        nome_grupo = identificador_grupo(
            grupo
        )

        for (
            dia_unidade,
            data_unidade,
        ) in unidades_temporais:

            variaveis = []

            for oferta in ofertas_grupo:
                for opcao in (
                    opcoes_por_oferta[
                        oferta.id
                    ]
                ):
                    if datas_provas is None:
                        pertence = (
                            opcao.dia
                            == dia_unidade
                        )
                    else:
                        pertence = (
                            opcao.data_prova
                            == data_unidade
                        )

                    if pertence:
                        variaveis.append(
                            x[
                                (
                                    oferta.id,
                                    opcao,
                                )
                            ]
                        )

            if not variaveis:
                continue

            max_provas = len(
                ofertas_grupo
            )

            if data_unidade is None:
                identificador = (
                    dia_unidade
                )
            else:
                identificador = (
                    data_unidade
                    .isoformat()
                    .replace("-", "_")
                )

            qtd = model.NewIntVar(
                0,
                max_provas,
                f"qtd_{nome_grupo}_{identificador}",
            )

            model.Add(
                qtd == sum(
                    variaveis
                )
            )

            # 2 ou mais provas.
            segunda = (
                model.NewBoolVar(
                    f"segunda_{nome_grupo}_{identificador}"
                )
            )

            model.Add(
                qtd >= 2
            ).OnlyEnforceIf(
                segunda
            )

            model.Add(
                qtd <= 1
            ).OnlyEnforceIf(
                segunda.Not()
            )

            termos_objetivo_alunos.append(
                PESO_SEGUNDA_PROVA
                * segunda
            )

            # 3 ou mais.
            terceira = (
                model.NewBoolVar(
                    f"terceira_{nome_grupo}_{identificador}"
                )
            )

            model.Add(
                qtd >= 3
            ).OnlyEnforceIf(
                terceira
            )

            model.Add(
                qtd <= 2
            ).OnlyEnforceIf(
                terceira.Not()
            )

            termos_objetivo_alunos.append(
                PESO_TERCEIRA_PROVA
                * terceira
            )

            # Cada prova acima da terceira.
            excesso = model.NewIntVar(
                0,
                max_provas,
                f"excesso_{nome_grupo}_{identificador}",
            )

            model.Add(
                excesso >= qtd - 3
            )

            termos_objetivo_alunos.append(
                PESO_QUARTA_OU_MAIS
                * excesso
            )

        # ----------------------------------------------------
        # MESMO DIA E MESMO HORÁRIO
        #
        # O bloco acima conta provas por DIA: duas provas no mesmo
        # dia são apenas indesejáveis. Duas provas no mesmo dia e
        # no mesmo horário são impossíveis para o aluno, então
        # entram no objetivo com um peso que domina todo o resto.
        # ----------------------------------------------------

        variaveis_por_faixa: dict[
            tuple[Any, str],
            list[Any],
        ] = {}

        for oferta in ofertas_grupo:
            for opcao in (
                opcoes_por_oferta[
                    oferta.id
                ]
            ):
                chave_faixa = (
                    (
                        opcao.dia
                        if datas_provas is None
                        else opcao.data_prova
                    ),
                    opcao.hora,
                )

                variaveis_por_faixa.setdefault(
                    chave_faixa,
                    [],
                ).append(
                    x[
                        (
                            oferta.id,
                            opcao,
                        )
                    ]
                )

        for indice_faixa, (
            _,
            variaveis_faixa,
        ) in enumerate(
            sorted(
                variaveis_por_faixa.items(),
                key=lambda item: (
                    str(item[0][0]),
                    item[0][1],
                ),
            )
        ):
            # Com uma opção só não há como haver sobreposição.
            if len(variaveis_faixa) < 2:
                continue

            sobreposicao = model.NewIntVar(
                0,
                len(variaveis_faixa) - 1,
                f"sobreposicao_{nome_grupo}_"
                f"{indice_faixa}",
            )

            model.Add(
                sobreposicao
                >= sum(variaveis_faixa) - 1
            )

            termos_objetivo_alunos.append(
                PESO_MESMO_HORARIO
                * sobreposicao
            )

    objetivo_alunos = sum(
        termos_objetivo_alunos
    )

    # --------------------------------------------------------
    # CONCENTRAÇÃO DAS PROVAS POR PROFESSOR
    #
    # Prioridades exatas da função objetivo:
    #   1. conflitos dos alunos (regra da versão 3);
    #   2. total de dias adicionais usados pelos professores;
    #   3. distância entre a primeira e a última prova de cada
    #      professor.
    #
    # O solver trabalha em duas etapas. Primeiro resolve exatamente o
    # mesmo objetivo da versão 3. Depois fixa esse valor e só então
    # otimiza os professores. Assim, nem um término por limite de tempo
    # na primeira etapa permite que a preferência secundária piore a
    # solução já encontrada para os alunos.
    # --------------------------------------------------------

    dias_adicionais_professores = []
    intervalos_professores = []
    professores_considerados = 0
    limite_intervalos = 0

    if concentrar_por_professor:
        ofertas_por_professor: dict[
            str,
            dict[int, OfertaAcademica],
        ] = {}

        for oferta in ofertas_ativas:
            for professor in oferta.professores:
                chave_professor = (
                    normalizar_nome_comparacao(
                        professor
                    )
                )

                if not chave_professor:
                    continue

                ofertas_por_professor.setdefault(
                    chave_professor,
                    {},
                )[oferta.id] = oferta

        if datas_provas is None:
            coordenada_temporal = {
                dia: ORDEM_DIAS[dia] - 1
                for dia, _
                in unidades_temporais
            }

            def chave_temporal(
                opcao: OpcaoProva,
            ) -> str | date:
                return opcao.dia

        else:
            data_base = min(
                datas_provas
            )

            coordenada_temporal = {
                data_real: (
                    data_real - data_base
                ).days
                for _, data_real
                in unidades_temporais
                if data_real is not None
            }

            def chave_temporal(
                opcao: OpcaoProva,
            ) -> str | date:
                if opcao.data_prova is None:
                    raise ValueError(
                        "Opção sem data no modo de datas reais."
                    )

                return opcao.data_prova

        coordenadas = list(
            coordenada_temporal.values()
        )

        menor_coordenada = min(
            coordenadas,
            default=0,
        )
        maior_coordenada = max(
            coordenadas,
            default=0,
        )
        intervalo_maximo = (
            maior_coordenada
            - menor_coordenada
        )

        for indice_professor, (
            _,
            ofertas_indexadas,
        ) in enumerate(
            sorted(
                ofertas_por_professor.items()
            )
        ):
            ofertas_professor = list(
                ofertas_indexadas.values()
            )

            # Um professor com uma única prova não possui o que
            # concentrar e não precisa gerar variáveis adicionais.
            if len(ofertas_professor) < 2:
                continue

            professores_considerados += 1
            usos_temporais = []

            for indice_unidade, (
                dia_unidade,
                data_unidade,
            ) in enumerate(
                unidades_temporais
            ):
                variaveis_unidade = []

                for oferta in ofertas_professor:
                    for opcao in (
                        opcoes_por_oferta[
                            oferta.id
                        ]
                    ):
                        if datas_provas is None:
                            pertence = (
                                opcao.dia
                                == dia_unidade
                            )
                        else:
                            pertence = (
                                opcao.data_prova
                                == data_unidade
                            )

                        if pertence:
                            variaveis_unidade.append(
                                x[
                                    (
                                        oferta.id,
                                        opcao,
                                    )
                                ]
                            )

                if not variaveis_unidade:
                    continue

                usou_unidade = model.NewBoolVar(
                    f"prof_{indice_professor}_"
                    f"usa_{indice_unidade}"
                )

                model.AddMaxEquality(
                    usou_unidade,
                    variaveis_unidade,
                )

                usos_temporais.append(
                    usou_unidade
                )

            max_dias_adicionais = max(
                0,
                min(
                    len(ofertas_professor),
                    len(usos_temporais),
                ) - 1,
            )

            dias_adicionais = model.NewIntVar(
                0,
                max_dias_adicionais,
                f"prof_{indice_professor}_dias_adicionais",
            )

            model.Add(
                dias_adicionais
                == sum(usos_temporais) - 1
            )

            dias_adicionais_professores.append(
                dias_adicionais
            )
            posicoes_provas = []

            for oferta in ofertas_professor:
                posicao = model.NewIntVar(
                    menor_coordenada,
                    maior_coordenada,
                    f"prof_{indice_professor}_"
                    f"oferta_{oferta.id}_posicao",
                )

                model.Add(
                    posicao
                    == sum(
                        coordenada_temporal[
                            chave_temporal(
                                opcao
                            )
                        ]
                        * x[
                            (
                                oferta.id,
                                opcao,
                            )
                        ]
                        for opcao in (
                            opcoes_por_oferta[
                                oferta.id
                            ]
                        )
                    )
                )

                posicoes_provas.append(
                    posicao
                )

            primeira_prova = model.NewIntVar(
                menor_coordenada,
                maior_coordenada,
                f"prof_{indice_professor}_primeira_prova",
            )
            ultima_prova = model.NewIntVar(
                menor_coordenada,
                maior_coordenada,
                f"prof_{indice_professor}_ultima_prova",
            )

            model.AddMinEquality(
                primeira_prova,
                posicoes_provas,
            )
            model.AddMaxEquality(
                ultima_prova,
                posicoes_provas,
            )

            intervalo = model.NewIntVar(
                0,
                intervalo_maximo,
                f"prof_{indice_professor}_intervalo",
            )

            model.Add(
                intervalo
                == ultima_prova - primeira_prova
            )

            intervalos_professores.append(
                intervalo
            )
            limite_intervalos += (
                intervalo_maximo
            )

    total_dias_adicionais = sum(
        dias_adicionais_professores
    )
    total_intervalos = sum(
        intervalos_professores
    )

    escala_dias_adicionais = (
        limite_intervalos + 1
    )

    objetivo_professores = (
        total_dias_adicionais
        * escala_dias_adicionais
        + total_intervalos
    )

    def criar_solver() -> cp_model.CpSolver:
        novo_solver = cp_model.CpSolver()

        novo_solver.parameters.max_time_in_seconds = (
            TEMPO_MAXIMO_SOLVER_SEGUNDOS
        )

        novo_solver.parameters.num_search_workers = (
            NUM_SEARCH_WORKERS
        )

        return novo_solver

    def valor_inteiro(
        solver_referencia: cp_model.CpSolver,
        expressao,
    ) -> int:
        if isinstance(expressao, int):
            return expressao

        return int(
            solver_referencia.Value(
                expressao
            )
        )

    # Etapa 1: objetivo original da versão 3.
    model.Minimize(
        objetivo_alunos
    )

    solver_alunos = criar_solver()
    status_alunos = solver_alunos.Solve(
        model
    )

    if status_alunos not in {
        cp_model.OPTIMAL,
        cp_model.FEASIBLE,
    }:
        raise RuntimeError(
            "Não foi encontrada uma solução viável."
        )

    objetivo_alunos_fixado = valor_inteiro(
        solver_alunos,
        objetivo_alunos,
    )

    solver = solver_alunos
    concentracao_otimizada = (
        concentrar_por_professor
        and professores_considerados == 0
    )

    # Etapa 2: preserva a solução dos alunos e concentra professores.
    if (
        concentrar_por_professor
        and professores_considerados > 0
    ):
        model.Add(
            objetivo_alunos
            == objetivo_alunos_fixado
        )

        model.Minimize(
            objetivo_professores
        )

        # A solução da primeira etapa já é uma solução viável para a
        # segunda; as dicas ajudam o solver a recuperá-la rapidamente.
        for variavel in x.values():
            model.AddHint(
                variavel,
                solver_alunos.Value(
                    variavel
                ),
            )

        solver_professores = criar_solver()
        status_professores = (
            solver_professores.Solve(
                model
            )
        )

        if status_professores in {
            cp_model.OPTIMAL,
            cp_model.FEASIBLE,
        }:
            solver = solver_professores
            concentracao_otimizada = True

    metricas = MetricasOtimizacao(
        objetivo_alunos=valor_inteiro(
            solver,
            objetivo_alunos
        ),
        dias_adicionais_professores=valor_inteiro(
            solver,
            total_dias_adicionais
        ),
        intervalo_total_professores=valor_inteiro(
            solver,
            total_intervalos
        ),
        professores_considerados=(
            professores_considerados
        ),
        concentracao_por_professor_ativa=(
            concentrar_por_professor
        ),
        concentracao_por_professor_otimizada=(
            concentracao_otimizada
        ),
    )

    escolhas: dict[
        int,
        OpcaoProva,
    ] = {}

    for oferta in ofertas_ativas:
        for opcao in (
            opcoes_por_oferta[
                oferta.id
            ]
        ):
            if solver.Value(
                x[
                    (
                        oferta.id,
                        opcao,
                    )
                ]
            ) == 1:
                escolhas[
                    oferta.id
                ] = opcao
                break

        if oferta.id not in escolhas:
            raise RuntimeError(
                "Não foi possível identificar a opção "
                f"escolhida para {oferta.codigo} / "
                f"turma {oferta.turma}."
            )

    return (
        solver,
        escolhas,
        opcoes_por_oferta,
        metricas,
    )


# ============================================================
# DATAFRAMES DE RESULTADO
# ============================================================

def criar_dataframe_calendario(
    ofertas: list[OfertaAcademica],
    escolhas: dict[int, OpcaoProva],
) -> pd.DataFrame:
    """
    Uma oferta compartilhada gera uma linha para cada grupo no
    DataFrame visual, mas todas essas linhas referem-se à MESMA
    prova escolhida.
    """
    colunas = [
        "curso",
        "periodo",
        "cod_turma",
        "data_prova",
        "dia_prova",
        "hora",
        "codigo",
        "materia",
        "turma",
        "sala",
        "professores",
        "oferta_compartilhada",
        "grupos_atendidos",
    ]

    linhas = []

    for oferta in ofertas:
        escolha = escolhas.get(
            oferta.id
        )

        if escolha is None:
            continue

        grupos_txt = " | ".join(
            rotulo_grupo(grupo)
            for grupo
            in sorted(
                oferta.grupos
            )
        )

        professores_txt = (
            " | ".join(
                sorted(
                    oferta.professores
                )
            )
        )

        # Mais de uma sala aqui significa divergência na grade para
        # a mesma prova; o texto mostra todas em vez de escolher uma.
        salas_txt = (
            " | ".join(
                sorted(
                    oferta.salas
                )
            )
            or SALA_INDEFINIDA
        )

        compartilhada = (
            "SIM"
            if len(
                oferta.grupos
            ) > 1
            else "NÃO"
        )

        for (
            curso,
            periodo,
            cod_turma,
        ) in sorted(
            oferta.grupos
        ):
            grupo_alunos = (
                curso,
                periodo,
                cod_turma,
            )

            linhas.append(
                {
                    "curso": curso,
                    "periodo": periodo,
                    "cod_turma": cod_turma,
                    "data_prova": (
                        formatar_data_br(
                            escolha.data_prova
                        )
                    ),
                    "dia_prova": (
                        escolha.dia
                    ),
                    "hora": (
                        escolha.hora
                    ),
                    "codigo": (
                        oferta.codigo
                    ),
                    "materia": (
                        oferta.materias_por_grupo.get(
                            grupo_alunos,
                            oferta.materia,
                        )
                    ),
                    "turma": (
                        oferta.turma
                    ),
                    "sala": (
                        salas_txt
                    ),
                    "professores": (
                        professores_txt
                    ),
                    "oferta_compartilhada": (
                        compartilhada
                    ),
                    "grupos_atendidos": (
                        grupos_txt
                    ),
                }
            )

    resultado = pd.DataFrame(
        linhas,
        columns=colunas,
    )

    if resultado.empty:
        return resultado

    if (
        resultado["data_prova"]
        .ne("")
        .any()
    ):
        resultado["_ordem"] = (
            pd.to_datetime(
                resultado[
                    "data_prova"
                ],
                format="%d/%m/%Y",
            )
        )
    else:
        resultado["_ordem"] = (
            resultado[
                "dia_prova"
            ]
            .map(
                ORDEM_DIAS
            )
        )

    return (
        resultado
        .sort_values(
            [
                "curso",
                "periodo",
                "cod_turma",
                "_ordem",
                "hora",
                "codigo",
            ]
        )
        .drop(
            columns="_ordem"
        )
        .reset_index(
            drop=True
        )
    )


def criar_dataframe_concentracao_professores(
    ofertas: list[OfertaAcademica],
    escolhas: dict[int, OpcaoProva],
) -> pd.DataFrame:
    """
    Resume a distribuição das provas sem duplicar ofertas
    compartilhadas entre cursos ou grupos de alunos.
    """
    colunas = [
        "professor",
        "quantidade_provas",
        "quantidade_dias",
        "intervalo_dias",
        "dias_utilizados",
        "provas",
        "situacao",
    ]

    por_professor: dict[
        str,
        dict[str, Any],
    ] = {}

    for oferta in ofertas:
        escolha = escolhas.get(
            oferta.id
        )

        if escolha is None:
            continue

        for professor in oferta.professores:
            chave = normalizar_nome_comparacao(
                professor
            )

            if not chave:
                continue

            registro = por_professor.setdefault(
                chave,
                {
                    "nomes": set(),
                    "provas": [],
                },
            )

            registro["nomes"].add(
                professor
            )
            registro["provas"].append(
                (
                    oferta,
                    escolha,
                )
            )

    linhas = []

    for chave in sorted(
        por_professor
    ):
        registro = por_professor[
            chave
        ]
        provas = registro[
            "provas"
        ]

        usa_datas_reais = any(
            escolha.data_prova is not None
            for _, escolha in provas
        )

        if usa_datas_reais:
            unidades = sorted(
                {
                    escolha.data_prova
                    for _, escolha in provas
                    if escolha.data_prova is not None
                }
            )

            intervalo_dias = (
                (unidades[-1] - unidades[0]).days
                if unidades
                else 0
            )

            rotulos_unidades = [
                f"{formatar_data_br(data_real)} "
                f"({nome_dia(data_real)})"
                for data_real in unidades
            ]

            def ordem_prova(
                item: tuple[
                    OfertaAcademica,
                    OpcaoProva,
                ],
            ):
                oferta, escolha = item
                return (
                    escolha.data_prova,
                    escolha.hora,
                    oferta.codigo,
                    oferta.turma,
                )

        else:
            unidades = sorted(
                {
                    escolha.dia
                    for _, escolha in provas
                },
                key=lambda dia: (
                    ORDEM_DIAS[dia]
                ),
            )

            intervalo_dias = (
                ORDEM_DIAS[unidades[-1]]
                - ORDEM_DIAS[unidades[0]]
                if unidades
                else 0
            )

            rotulos_unidades = list(
                unidades
            )

            def ordem_prova(
                item: tuple[
                    OfertaAcademica,
                    OpcaoProva,
                ],
            ):
                oferta, escolha = item
                return (
                    ORDEM_DIAS[
                        escolha.dia
                    ],
                    escolha.hora,
                    oferta.codigo,
                    oferta.turma,
                )

        provas_ordenadas = sorted(
            provas,
            key=ordem_prova,
        )

        detalhes_provas = []

        for oferta, escolha in (
            provas_ordenadas
        ):
            if escolha.data_prova is None:
                quando = (
                    f"{escolha.dia} "
                    f"{escolha.hora}"
                )
            else:
                quando = (
                    f"{formatar_data_br(escolha.data_prova)} "
                    f"{escolha.hora}"
                )

            detalhes_provas.append(
                f"{oferta.codigo} / turma {oferta.turma} — "
                f"{texto_materias_oferta(oferta)} — {quando}"
            )

        quantidade_dias = len(
            unidades
        )

        linhas.append(
            {
                "professor": sorted(
                    registro["nomes"],
                    key=str.casefold,
                )[0],
                "quantidade_provas": len(
                    provas
                ),
                "quantidade_dias": (
                    quantidade_dias
                ),
                "intervalo_dias": (
                    intervalo_dias
                ),
                "dias_utilizados": " | ".join(
                    rotulos_unidades
                ),
                "provas": "\n".join(
                    detalhes_provas
                ),
                "situacao": (
                    "CONCENTRADO EM 1 DIA"
                    if quantidade_dias == 1
                    else f"DISTRIBUÍDO EM {quantidade_dias} DIAS"
                ),
            }
        )

    if not linhas:
        return pd.DataFrame(
            columns=colunas
        )

    return (
        pd.DataFrame(
            linhas,
            columns=colunas,
        )
        .sort_values(
            [
                "quantidade_dias",
                "intervalo_dias",
                "professor",
            ],
            ascending=[
                False,
                False,
                True,
            ],
        )
        .reset_index(
            drop=True
        )
    )


def criar_dataframe_ofertas(
    ofertas: list[OfertaAcademica],
    escolhas: dict[int, OpcaoProva],
    opcoes_por_oferta: dict[
        int,
        list[OpcaoProva],
    ],
) -> pd.DataFrame:
    linhas = []

    for oferta in ofertas:
        escolha = escolhas.get(
            oferta.id
        )

        opcoes_grade = sorted(
            oferta.opcoes_semanais,
            key=lambda o: (
                ORDEM_DIAS[
                    o.dia
                ],
                o.hora,
            ),
        )

        opcoes_elegiveis = (
            opcoes_por_oferta.get(
                oferta.id,
                [],
            )
        )

        linhas.append(
            {
                "codigo": oferta.codigo,
                "materia": texto_materias_oferta(
                    oferta
                ),
                "turma": oferta.turma,
                "professores": (
                    " | ".join(
                        sorted(
                            oferta.professores
                        )
                    )
                ),
                "grupos_atendidos": (
                    " | ".join(
                        rotulo_grupo(grupo)
                        for grupo
                        in sorted(
                            oferta.grupos
                        )
                    )
                ),
                "quantidade_grupos": (
                    len(
                        oferta.grupos
                    )
                ),
                "oferta_compartilhada": (
                    "SIM"
                    if len(
                        oferta.grupos
                    ) > 1
                    else "NÃO"
                ),
                "opcoes_da_grade": (
                    " | ".join(
                        f"{o.dia} {o.hora}"
                        for o in opcoes_grade
                    )
                ),
                "opcoes_elegiveis": (
                    " | ".join(
                        (
                            f"{formatar_data_br(o.data_prova)} "
                            f"{o.dia} {o.hora}"
                            if o.data_prova
                            else f"{o.dia} {o.hora}"
                        )
                        for o
                        in opcoes_elegiveis
                    )
                ),
                "status": (
                    "ALOCADA"
                    if escolha is not None
                    else (
                        "NÃO ALOCADA — "
                        "TRATAR MANUALMENTE"
                    )
                ),
                "data_escolhida": (
                    formatar_data_br(
                        escolha.data_prova
                    )
                    if escolha
                    else ""
                ),
                "dia_escolhido": (
                    escolha.dia
                    if escolha
                    else ""
                ),
                "hora_escolhida": (
                    escolha.hora
                    if escolha
                    else ""
                ),
            }
        )

    return (
        pd.DataFrame(
            linhas
        )
        .sort_values(
            [
                "codigo",
                "turma",
            ]
        )
        .reset_index(
            drop=True
        )
    )


def criar_dataframe_nao_alocadas(
    ofertas: list[OfertaAcademica],
    opcoes_por_oferta: dict[
        int,
        list[OpcaoProva],
    ],
    datas_provas: list[date] | None,
) -> pd.DataFrame:
    colunas = [
        "codigo",
        "materia",
        "turma",
        "grupos_atendidos",
        "professores",
        "dias_da_grade",
        "datas_disponiveis",
        "motivo",
    ]

    linhas = []

    for oferta in ofertas:
        opcoes = (
            opcoes_por_oferta.get(
                oferta.id,
                [],
            )
        )

        if opcoes:
            continue

        dias_grade = sorted(
            {
                opcao.dia
                for opcao
                in oferta.opcoes_semanais
            },
            key=lambda dia: (
                ORDEM_DIAS[dia]
            ),
        )

        if datas_provas:
            datas_txt = " | ".join(
                f"{formatar_data_br(data_real)} "
                f"({nome_dia(data_real)})"
                for data_real
                in datas_provas
            )

            motivo = (
                "Nenhuma data permitida coincide com os "
                "dias de aula desta oferta. Tratar manualmente."
            )
        else:
            datas_txt = ""
            motivo = (
                "Nenhuma opção elegível foi encontrada. "
                "Tratar manualmente."
            )

        linhas.append(
            {
                "codigo": oferta.codigo,
                "materia": texto_materias_oferta(
                    oferta
                ),
                "turma": oferta.turma,
                "grupos_atendidos": (
                    " | ".join(
                        rotulo_grupo(grupo)
                        for grupo
                        in sorted(
                            oferta.grupos
                        )
                    )
                ),
                "professores": (
                    " | ".join(
                        sorted(
                            oferta.professores
                        )
                    )
                ),
                "dias_da_grade": (
                    " / ".join(
                        dias_grade
                    )
                ),
                "datas_disponiveis": (
                    datas_txt
                ),
                "motivo": motivo,
            }
        )

    return pd.DataFrame(
        linhas,
        columns=colunas,
    )


def criar_dataframe_conflitos(
    calendario: pd.DataFrame,
) -> pd.DataFrame:
    colunas = [
        "curso",
        "periodo",
        "cod_turma",
        "data",
        "dia",
        "quantidade_provas",
        "gravidade",
        "choque_mesmo_horario",
        "horarios_com_choque",
        "ofertas",
    ]

    if calendario.empty:
        return pd.DataFrame(
            columns=colunas
        )

    usa_datas = (
        calendario[
            "data_prova"
        ]
        .ne("")
        .any()
    )

    # O choque só existe dentro de uma mesma turma de alunos:
    # curso + período + cod_turma.
    chaves = [
        "curso",
        "periodo",
        "cod_turma",
        (
            "data_prova"
            if usa_datas
            else "dia_prova"
        ),
    ]

    linhas = []

    for chave, grupo in (
        calendario.groupby(
            chaves
        )
    ):
        curso = chave[0]
        periodo = chave[1]
        cod_turma = chave[2]
        unidade = chave[3]

        quantidade = len(
            grupo
        )

        if quantidade <= 1:
            continue

        dia = (
            grupo[
                "dia_prova"
            ]
            .iloc[0]
        )

        contagem_hora = (
            grupo.groupby(
                "hora"
            )
            .size()
        )

        horarios_choque = [
            str(horario)
            for horario, qtd
            in contagem_hora.items()
            if qtd > 1
        ]

        if quantidade == 2:
            gravidade = (
                "Duas provas no mesmo dia"
            )
        elif quantidade == 3:
            gravidade = (
                "TRÊS provas no mesmo dia"
            )
        else:
            gravidade = (
                f"{quantidade} provas no mesmo dia"
            )

        ofertas_txt = " | ".join(
            (
                f"{row['codigo']} - "
                f"{row['materia']} "
                f"(turma {row['turma']})"
            )
            for _, row
            in grupo.iterrows()
        )

        linhas.append(
            {
                "curso": curso,
                "periodo": periodo,
                "cod_turma": cod_turma,
                "data": (
                    unidade
                    if usa_datas
                    else ""
                ),
                "dia": dia,
                "quantidade_provas": (
                    quantidade
                ),
                "gravidade": (
                    gravidade
                ),
                "choque_mesmo_horario": (
                    "SIM"
                    if horarios_choque
                    else "NÃO"
                ),
                "horarios_com_choque": (
                    ", ".join(
                        horarios_choque
                    )
                ),
                "ofertas": (
                    ofertas_txt
                ),
            }
        )

    return pd.DataFrame(
        linhas,
        columns=colunas,
    )


def criar_dataframe_diagnostico(
    calendario: pd.DataFrame,
) -> pd.DataFrame:
    colunas = [
        "curso",
        "periodo",
        "cod_turma",
        "numero_ofertas",
        "dias_com_prova",
        "dias_com_conflito",
        "max_provas_mesmo_dia",
        "situacao",
    ]

    if calendario.empty:
        return pd.DataFrame(
            columns=colunas
        )

    usa_datas = (
        calendario[
            "data_prova"
        ]
        .ne("")
        .any()
    )

    campo_temporal = (
        "data_prova"
        if usa_datas
        else "dia_prova"
    )

    linhas = []

    for (
        curso,
        periodo,
        cod_turma,
    ), grupo in calendario.groupby(
        [
            "curso",
            "periodo",
            "cod_turma",
        ]
    ):
        contagem = (
            grupo.groupby(
                campo_temporal
            )
            .size()
        )

        conflitos = int(
            (
                contagem > 1
            ).sum()
        )

        linhas.append(
            {
                "curso": curso,
                "periodo": periodo,
                "cod_turma": cod_turma,
                "numero_ofertas": (
                    len(grupo)
                ),
                "dias_com_prova": (
                    len(contagem)
                ),
                "dias_com_conflito": (
                    conflitos
                ),
                "max_provas_mesmo_dia": (
                    int(
                        contagem.max()
                    )
                ),
                "situacao": (
                    "SEM CONFLITOS"
                    if conflitos == 0
                    else "COM CONFLITOS"
                ),
            }
        )

    return (
        pd.DataFrame(
            linhas
        )
        .sort_values(
            [
                "curso",
                "periodo",
                "cod_turma",
            ]
        )
        .reset_index(
            drop=True
        )
    )


def criar_dataframe_parametros(
    datas_provas: list[date] | None,
    nome_avaliacao: str | None,
    concentrar_por_professor: bool,
) -> pd.DataFrame:
    linhas = [
        {
            "parametro": "versão_do_algoritmo",
            "valor": VERSAO_ALGORITMO,
        },
        {
            "parametro": "avaliação",
            "valor": (
                nome_avaliacao
                or ""
            ),
        },
        {
            "parametro": "modo",
            "valor": (
                "datas reais"
                if datas_provas
                else "dias da semana"
            ),
        },
        {
            "parametro": "chave_da_oferta",
            "valor": (
                "codigo + turma + agenda + professor normalizado; "
                "o nome da disciplina não participa da identidade"
            ),
        },
        {
            "parametro": "grupo_de_alunos",
            "valor": (
                "curso + periodo + cod_turma "
                "(turmas diferentes do mesmo período são "
                "grupos independentes)"
            ),
        },
        {
            "parametro": "restricao_principal",
            "valor": (
                "minimizar mais de uma prova no mesmo dia "
                "para cada curso + período + turma"
            ),
        },
        {
            "parametro": "concentração_por_professor",
            "valor": (
                "ativada"
                if concentrar_por_professor
                else "desativada"
            ),
        },
        {
            "parametro": "prioridades_da_otimização",
            "valor": (
                "1) conflitos dos alunos; "
                "2) quantidade de dias por professor; "
                "3) intervalo entre as provas do professor"
            ),
        },
        {
            "parametro": "ofertas_compartilhadas",
            "valor": (
                "uma única prova; a mesma decisão vale para "
                "todos os cursos/períodos atendidos"
            ),
        },
    ]

    if datas_provas:
        for indice, data_real in enumerate(
            datas_provas,
            start=1,
        ):
            linhas.append(
                {
                    "parametro": (
                        f"data_disponivel_{indice}"
                    ),
                    "valor": (
                        f"{formatar_data_br(data_real)} "
                        f"({nome_dia(data_real)})"
                    ),
                }
            )

    return pd.DataFrame(
        linhas
    )


# ============================================================
# EXCEL — ESTILO
# ============================================================

def ajustar_larguras_aba(
    ws,
    largura_maxima: int = 60,
) -> None:
    for coluna in ws.columns:
        try:
            indice = (
                coluna[0].column
            )
        except AttributeError:
            continue

        letra = (
            get_column_letter(
                indice
            )
        )

        maior = 0

        for celula in coluna:
            if celula.value is None:
                continue

            linhas = (
                str(celula.value)
                .splitlines()
            )

            maior = max(
                maior,
                max(
                    (
                        len(linha)
                        for linha in linhas
                    ),
                    default=0,
                ),
            )

        ws.column_dimensions[
            letra
        ].width = min(
            max(
                maior + 2,
                10,
            ),
            largura_maxima,
        )


def formatar_aba_tabela(
    ws,
) -> None:
    fill = PatternFill(
        "solid",
        fgColor="17365D",
    )

    fonte = Font(
        color="FFFFFF",
        bold=True,
    )

    linha_fina = Side(
        style="thin",
        color="D9D9D9",
    )

    for celula in ws[1]:
        celula.fill = fill
        celula.font = fonte
        celula.alignment = (
            Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        )

    for linha in ws.iter_rows(
        min_row=2
    ):
        for celula in linha:
            celula.border = Border(
                bottom=linha_fina
            )

            celula.alignment = (
                Alignment(
                    vertical="top",
                    wrap_text=True,
                )
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        ws.dimensions
    )
    ws.sheet_view.showGridLines = (
        False
    )

    ajustar_larguras_aba(
        ws
    )


# ============================================================
# EXCEL — CALENDÁRIO VISUAL
# ============================================================

def criar_aba_calendario_visual(
    writer,
    calendario: pd.DataFrame,
    datas_provas: list[date] | None,
    nome_avaliacao: str | None,
    nao_alocadas: pd.DataFrame,
    excluidas: pd.DataFrame,
) -> None:
    wb = writer.book

    if (
        "Calendário Visual"
        in wb.sheetnames
    ):
        del wb[
            "Calendário Visual"
        ]

    ws = wb.create_sheet(
        "Calendário Visual",
        0,
    )

    fill_titulo = PatternFill(
        "solid",
        fgColor="17365D",
    )

    fill_cabecalho = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    fill_dia_conflito = PatternFill(
        "solid",
        fgColor="FFD966",
    )

    fill_conflito_horario = (
        PatternFill(
            "solid",
            fgColor="FFC7CE",
        )
    )

    fill_prova = PatternFill(
        "solid",
        fgColor="E2F0D9",
    )

    fill_compartilhada = (
        PatternFill(
            "solid",
            fgColor="E4DFEC",
        )
    )

    fonte_titulo = Font(
        color="FFFFFF",
        bold=True,
        size=14,
    )

    fonte_cabecalho = Font(
        bold=True,
    )

    fonte_alerta = Font(
        bold=True,
        color="9C6500",
    )

    fina = Side(
        style="thin",
        color="A6A6A6",
    )

    borda = Border(
        left=fina,
        right=fina,
        top=fina,
        bottom=fina,
    )

    alinhamento = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    ws.sheet_view.showGridLines = (
        False
    )
    ws.page_setup.orientation = (
        "landscape"
    )
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = (
        True
    )

    # --------------------------------------------------------
    # COLUNAS TEMPORAIS
    # --------------------------------------------------------

    if datas_provas:
        # Mesma ordem do layout oficial: segunda a sexta.
        colunas_temporais = [
            {
                "chave": (
                    formatar_data_br(
                        data_real
                    )
                ),
                "cabecalho": (
                    f"{data_real.strftime('%d/%m')}\n"
                    f"{nome_dia(data_real)}"
                ),
            }
            for data_real
            in ordenar_datas_por_dia_semana(
                datas_provas
            )
        ]
    else:
        colunas_temporais = [
            {
                "chave": dia,
                "cabecalho": dia,
            }
            for dia
            in DIAS_VISUAIS_PADRAO
        ]

    numero_colunas = (
        1
        + len(
            colunas_temporais
        )
    )

    ws.column_dimensions[
        "A"
    ].width = 12

    for coluna in range(
        2,
        numero_colunas + 1,
    ):
        ws.column_dimensions[
            get_column_letter(
                coluna
            )
        ].width = 36

    linha_atual = 1

    # --------------------------------------------------------
    # CALENDÁRIOS POR CURSO + PERÍODO + TURMA
    # --------------------------------------------------------

    if not calendario.empty:
        grupos = (
            calendario.groupby(
                [
                    "curso",
                    "periodo",
                    "cod_turma",
                ],
                sort=True,
            )
        )

        for (
            curso,
            periodo,
            cod_turma,
        ), grupo in grupos:

            ws.merge_cells(
                start_row=linha_atual,
                start_column=1,
                end_row=linha_atual,
                end_column=numero_colunas,
            )

            titulo = ws.cell(
                row=linha_atual,
                column=1,
            )

            prefixo = (
                f"{nome_avaliacao} — "
                if nome_avaliacao
                else ""
            )

            sufixo_turma = (
                f" — TURMA {cod_turma}"
                if cod_turma
                else ""
            )

            titulo.value = (
                f"{prefixo}"
                f"{curso} — "
                f"{periodo}º PERÍODO"
                f"{sufixo_turma}"
            )

            titulo.fill = (
                fill_titulo
            )
            titulo.font = (
                fonte_titulo
            )
            titulo.alignment = (
                alinhamento
            )

            for coluna in range(
                1,
                numero_colunas + 1,
            ):
                ws.cell(
                    row=linha_atual,
                    column=coluna,
                ).fill = fill_titulo

            ws.row_dimensions[
                linha_atual
            ].height = 28

            linha_atual += 1

            if datas_provas:
                contagem = (
                    grupo.groupby(
                        "data_prova"
                    )
                    .size()
                    .to_dict()
                )
            else:
                contagem = (
                    grupo.groupby(
                        "dia_prova"
                    )
                    .size()
                    .to_dict()
                )

            # Cabeçalho.
            celula = ws.cell(
                row=linha_atual,
                column=1,
                value="Horário",
            )

            celula.fill = (
                fill_cabecalho
            )
            celula.font = (
                fonte_cabecalho
            )
            celula.border = borda
            celula.alignment = (
                alinhamento
            )

            for indice, info in enumerate(
                colunas_temporais,
                start=2,
            ):
                qtd = contagem.get(
                    info["chave"],
                    0,
                )

                celula = ws.cell(
                    row=linha_atual,
                    column=indice,
                )

                if qtd > 1:
                    celula.value = (
                        f"{info['cabecalho']}\n"
                        f"⚠ {qtd} PROVAS"
                    )
                    celula.fill = (
                        fill_dia_conflito
                    )
                    celula.font = (
                        fonte_alerta
                    )
                else:
                    celula.value = (
                        info["cabecalho"]
                    )
                    celula.fill = (
                        fill_cabecalho
                    )
                    celula.font = (
                        fonte_cabecalho
                    )

                celula.border = borda
                celula.alignment = (
                    alinhamento
                )

            ws.row_dimensions[
                linha_atual
            ].height = 42

            linha_atual += 1

            horarios_encontrados = set(
                grupo["hora"]
                .dropna()
                .astype(str)
            )

            horarios = list(
                ORDEM_HORARIOS
            )

            horarios.extend(
                sorted(
                    horarios_encontrados
                    - set(
                        ORDEM_HORARIOS
                    )
                )
            )

            for horario in horarios:
                celula_hora = ws.cell(
                    row=linha_atual,
                    column=1,
                    value=horario,
                )

                celula_hora.font = (
                    fonte_cabecalho
                )
                celula_hora.fill = (
                    fill_cabecalho
                )
                celula_hora.border = (
                    borda
                )
                celula_hora.alignment = (
                    alinhamento
                )

                for indice, info in enumerate(
                    colunas_temporais,
                    start=2,
                ):
                    if datas_provas:
                        mascara_tempo = (
                            grupo[
                                "data_prova"
                            ]
                            == info["chave"]
                        )
                    else:
                        mascara_tempo = (
                            grupo[
                                "dia_prova"
                            ]
                            == info["chave"]
                        )

                    provas = grupo[
                        mascara_tempo
                        & (
                            grupo["hora"]
                            == horario
                        )
                    ]

                    celula = ws.cell(
                        row=linha_atual,
                        column=indice,
                    )

                    if provas.empty:
                        celula.value = ""

                    else:
                        textos = []

                        for _, prova in (
                            provas.iterrows()
                        ):
                            texto = (
                                f"{prova['codigo']} — "
                                f"{prova['materia']}\n"
                                f"Turma {prova['turma']}\n"
                                f"{prova['professores']}"
                            )

                            if (
                                prova[
                                    "oferta_compartilhada"
                                ]
                                == "SIM"
                            ):
                                texto += (
                                    "\n"
                                    "Oferta compartilhada: "
                                    f"{prova['grupos_atendidos']}"
                                )

                            textos.append(
                                texto
                            )

                        celula.value = (
                            "\n\n".join(
                                textos
                            )
                        )

                        if len(provas) > 1:
                            celula.fill = (
                                fill_conflito_horario
                            )
                        elif (
                            provas[
                                "oferta_compartilhada"
                            ]
                            .eq("SIM")
                            .any()
                        ):
                            celula.fill = (
                                fill_compartilhada
                            )
                        else:
                            celula.fill = (
                                fill_prova
                            )

                    celula.border = (
                        borda
                    )
                    celula.alignment = (
                        alinhamento
                    )

                ws.row_dimensions[
                    linha_atual
                ].height = 92

                linha_atual += 1

            linha_atual += 1

            ws.cell(
                row=linha_atual,
                column=1,
                value="Legenda:",
            ).font = Font(
                bold=True
            )

            legendas = [
                (
                    "Prova",
                    fill_prova,
                ),
                (
                    "Oferta compartilhada",
                    fill_compartilhada,
                ),
                (
                    "Dia com 2+ provas",
                    fill_dia_conflito,
                ),
                (
                    "Choque no mesmo horário",
                    fill_conflito_horario,
                ),
            ]

            for coluna, (
                texto,
                fill,
            ) in enumerate(
                legendas,
                start=2,
            ):
                if coluna > numero_colunas:
                    break

                c = ws.cell(
                    row=linha_atual,
                    column=coluna,
                    value=texto,
                )
                c.fill = fill
                c.alignment = (
                    alinhamento
                )

            linha_atual += 3

    # --------------------------------------------------------
    # NÃO ALOCADAS
    # --------------------------------------------------------

    if not nao_alocadas.empty:
        largura = max(
            numero_colunas,
            8,
        )

        ws.merge_cells(
            start_row=linha_atual,
            start_column=1,
            end_row=linha_atual,
            end_column=largura,
        )

        c = ws.cell(
            row=linha_atual,
            column=1,
            value=(
                "OFERTAS NÃO ALOCADAS AUTOMATICAMENTE — "
                "TRATAR MANUALMENTE"
            ),
        )

        c.fill = (
            fill_dia_conflito
        )
        c.font = Font(
            bold=True,
            size=12,
        )
        c.alignment = (
            alinhamento
        )

        linha_atual += 1

        cabecalhos = [
            "Código",
            "Matéria",
            "Turma",
            "Grupos",
            "Professores",
            "Dias da grade",
            "Datas disponíveis",
            "Motivo",
        ]

        for coluna, cabecalho in enumerate(
            cabecalhos,
            start=1,
        ):
            c = ws.cell(
                row=linha_atual,
                column=coluna,
                value=cabecalho,
            )
            c.fill = (
                fill_cabecalho
            )
            c.font = (
                fonte_cabecalho
            )
            c.border = borda
            c.alignment = (
                alinhamento
            )

        linha_atual += 1

        for _, row in (
            nao_alocadas.iterrows()
        ):
            valores = [
                row["codigo"],
                row["materia"],
                row["turma"],
                row[
                    "grupos_atendidos"
                ],
                row["professores"],
                row["dias_da_grade"],
                row[
                    "datas_disponiveis"
                ],
                row["motivo"],
            ]

            for coluna, valor in enumerate(
                valores,
                start=1,
            ):
                c = ws.cell(
                    row=linha_atual,
                    column=coluna,
                    value=valor,
                )
                c.border = borda
                c.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            ws.row_dimensions[
                linha_atual
            ].height = 55

            linha_atual += 1

        linha_atual += 2

    # --------------------------------------------------------
    # ED / ELETIVAS
    # --------------------------------------------------------

    if not excluidas.empty:
        largura = max(
            numero_colunas,
            7,
        )

        ws.merge_cells(
            start_row=linha_atual,
            start_column=1,
            end_row=linha_atual,
            end_column=largura,
        )

        c = ws.cell(
            row=linha_atual,
            column=1,
            value=(
                "OFERTAS EXCLUÍDAS DA ALOCAÇÃO — "
                "ESTUDO DIRIGIDO (ED) / ELETIVAS (E)"
            ),
        )

        c.fill = PatternFill(
            "solid",
            fgColor="D9D9D9",
        )
        c.font = Font(
            bold=True,
            size=12,
        )
        c.alignment = (
            alinhamento
        )

        linha_atual += 1

        cabecalhos = [
            "Código",
            "Matéria",
            "Turma",
            "Grupos",
            "Professores",
            "Tipo",
            "Motivo",
        ]

        for coluna, cabecalho in enumerate(
            cabecalhos,
            start=1,
        ):
            c = ws.cell(
                row=linha_atual,
                column=coluna,
                value=cabecalho,
            )
            c.fill = (
                fill_cabecalho
            )
            c.font = (
                fonte_cabecalho
            )
            c.border = borda
            c.alignment = (
                alinhamento
            )

        linha_atual += 1

        for _, row in (
            excluidas.iterrows()
        ):
            valores = [
                row["codigo"],
                row["materia"],
                row["turma"],
                row[
                    "grupos_atendidos"
                ],
                row["professores"],
                row["tipo"],
                row["motivo"],
            ]

            for coluna, valor in enumerate(
                valores,
                start=1,
            ):
                c = ws.cell(
                    row=linha_atual,
                    column=coluna,
                    value=valor,
                )
                c.border = borda
                c.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            ws.row_dimensions[
                linha_atual
            ].height = 50

            linha_atual += 1


# ============================================================
# EXCEL — EXPORTADOR PROVISÓRIO DA PRIMEIRA REVISÃO 4.2
# Mantido apenas como referência interna; a função pública utilizada
# pelo programa é redefinida adiante com o layout institucional.
# ============================================================

def nome_seguro_arquivo(
    valor: Any,
) -> str:
    """Converte um rótulo em uma parte segura de nome de arquivo."""
    texto = unicodedata.normalize(
        "NFKD",
        limpar_texto(valor),
    ).encode(
        "ascii",
        "ignore",
    ).decode(
        "ascii"
    )

    partes = []
    anterior_foi_separador = False

    for caractere in texto:
        if caractere.isalnum():
            partes.append(caractere)
            anterior_foi_separador = False
        elif not anterior_foi_separador:
            partes.append("_")
            anterior_foi_separador = True

    return (
        "".join(partes)
        .strip("_")
        or "calendario"
    )


def materia_sem_codigo(
    materia: Any,
    codigo: Any,
) -> str:
    """Evita repetir o código quando ele já está no fim da matéria."""
    nome = limpar_texto(
        materia
    )
    codigo_txt = limpar_texto(
        codigo
    )

    if (
        codigo_txt
        and nome.casefold().endswith(
            codigo_txt.casefold()
        )
    ):
        nome = nome[
            :-len(codigo_txt)
        ].rstrip(
            " -–—"
        )

    return nome


def criar_arquivo_modelo_oficial(
    curso: str,
    nome_avaliacao: str,
    periodo_letivo: str,
    datas_provas: list[date],
) -> tuple[Workbook, Any, int]:
    """
    Cria em memória o arquivo-modelo oficial com uma coluna por data.

    A identidade da coluna é a data completa, e não o nome do dia da
    semana. Portanto, duas quartas-feiras ou duas quintas-feiras são
    mantidas como colunas independentes, em ordem cronológica.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendário"

    quantidade_colunas = (
        1
        + len(datas_provas)
    )

    cor_ibmec = "C8102E"
    cor_cabecalho = "E7E6E6"
    branca = "FFFFFF"

    fill_titulo = PatternFill(
        "solid",
        fgColor=cor_ibmec,
    )
    fill_cabecalho = PatternFill(
        "solid",
        fgColor=cor_cabecalho,
    )

    alinhamento = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=quantidade_colunas,
    )
    ws.cell(
        row=1,
        column=1,
        value="IBMEC — CALENDÁRIO DE PROVAS",
    )

    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=quantidade_colunas,
    )
    ws.cell(
        row=2,
        column=1,
        value=(
            f"{nome_avaliacao} — "
            f"{periodo_letivo} — "
            f"{curso}"
        ),
    )

    for linha in (1, 2):
        celula = ws.cell(
            row=linha,
            column=1,
        )
        celula.fill = fill_titulo
        celula.font = Font(
            color=branca,
            bold=True,
            size=(
                15
                if linha == 1
                else 12
            ),
        )
        celula.alignment = alinhamento

        for coluna in range(
            1,
            quantidade_colunas + 1,
        ):
            ws.cell(
                row=linha,
                column=coluna,
            ).fill = fill_titulo

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    ws.column_dimensions["A"].width = 12

    for coluna in range(
        2,
        quantidade_colunas + 1,
    ):
        ws.column_dimensions[
            get_column_letter(coluna)
        ].width = 31

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # A primeira seção começa na linha 4. O cabeçalho cinza será
    # aplicado por seção, pois um curso pode ter vários períodos.
    ws.cell(
        row=3,
        column=1,
    ).fill = fill_cabecalho

    return wb, ws, 4


def preencher_bloco_modelo_oficial(
    ws,
    linha_inicial: int,
    grupo: pd.DataFrame,
    periodo: int,
    cod_turma: str,
    datas_provas: list[date],
) -> int:
    """Preenche um bloco de período/turma e devolve a próxima linha."""
    quantidade_colunas = (
        1
        + len(datas_provas)
    )

    fill_periodo = PatternFill(
        "solid",
        fgColor="404040",
    )
    fill_cabecalho = PatternFill(
        "solid",
        fgColor="D9E1F2",
    )
    fill_prova = PatternFill(
        "solid",
        fgColor="FFFFFF",
    )
    fill_conflito = PatternFill(
        "solid",
        fgColor="FFC7CE",
    )
    fill_horario = PatternFill(
        "solid",
        fgColor="E7E6E6",
    )

    fina = Side(
        style="thin",
        color="A6A6A6",
    )
    borda = Border(
        left=fina,
        right=fina,
        top=fina,
        bottom=fina,
    )
    alinhamento = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    ws.merge_cells(
        start_row=linha_inicial,
        start_column=1,
        end_row=linha_inicial,
        end_column=quantidade_colunas,
    )

    turma_txt = (
        f" — TURMA {cod_turma}"
        if cod_turma
        else ""
    )

    titulo = ws.cell(
        row=linha_inicial,
        column=1,
        value=(
            f"{periodo}º PERÍODO"
            f"{turma_txt}"
        ),
    )
    titulo.font = Font(
        color="FFFFFF",
        bold=True,
        size=12,
    )
    titulo.alignment = alinhamento

    for coluna in range(
        1,
        quantidade_colunas + 1,
    ):
        ws.cell(
            row=linha_inicial,
            column=coluna,
        ).fill = fill_periodo

    ws.row_dimensions[
        linha_inicial
    ].height = 24

    linha_cabecalho = (
        linha_inicial + 1
    )

    cabecalhos = [
        "Horário",
        *[
            (
                f"{nome_dia(data_real)}\n"
                f"{data_real.strftime('%d/%m/%Y')}"
            )
            for data_real
            in datas_provas
        ],
    ]

    for coluna, cabecalho in enumerate(
        cabecalhos,
        start=1,
    ):
        celula = ws.cell(
            row=linha_cabecalho,
            column=coluna,
            value=cabecalho,
        )
        celula.fill = fill_cabecalho
        celula.font = Font(
            bold=True
        )
        celula.border = borda
        celula.alignment = alinhamento

    ws.row_dimensions[
        linha_cabecalho
    ].height = 38

    horarios_encontrados = set(
        grupo["hora"]
        .dropna()
        .astype(str)
    )
    horarios = list(
        ORDEM_HORARIOS
    )
    horarios.extend(
        sorted(
            horarios_encontrados
            - set(ORDEM_HORARIOS)
        )
    )

    linha_atual = (
        linha_cabecalho + 1
    )

    for horario in horarios:
        celula_hora = ws.cell(
            row=linha_atual,
            column=1,
            value=horario,
        )
        celula_hora.fill = fill_horario
        celula_hora.font = Font(
            bold=True
        )
        celula_hora.border = borda
        celula_hora.alignment = alinhamento

        maior_quantidade = 1

        for indice, data_real in enumerate(
            datas_provas,
            start=2,
        ):
            chave_data = formatar_data_br(
                data_real
            )
            provas = grupo[
                (
                    grupo["data_prova"]
                    == chave_data
                )
                & (
                    grupo["hora"]
                    == horario
                )
            ]

            celula = ws.cell(
                row=linha_atual,
                column=indice,
            )
            celula.border = borda
            celula.alignment = alinhamento

            if provas.empty:
                celula.value = ""
                celula.fill = fill_prova
                continue

            textos = []

            for _, prova in provas.iterrows():
                nome_materia = materia_sem_codigo(
                    prova["materia"],
                    prova["codigo"],
                )
                sala = (
                    limpar_texto(
                        prova["sala"]
                    )
                    or SALA_INDEFINIDA
                )

                textos.append(
                    (
                        f"{nome_materia}\n"
                        f"{prova['codigo']}\n"
                        f"Sala {sala}    "
                        f"Turma {prova['turma']}\n"
                        f"{prova['professores']}"
                    )
                )

            celula.value = "\n\n".join(
                textos
            )
            celula.fill = (
                fill_conflito
                if len(provas) > 1
                else fill_prova
            )
            maior_quantidade = max(
                maior_quantidade,
                len(provas),
            )

        ws.row_dimensions[
            linha_atual
        ].height = max(
            76,
            76 * maior_quantidade,
        )

        linha_atual += 1

    return linha_atual + 2


def exportar_calendarios_oficiais_provisorio(
    calendario: pd.DataFrame,
    nome_avaliacao: str,
    periodo_letivo: str,
    datas_provas: Iterable[str | date | datetime] | None,
    pasta_saida: str | Path,
) -> list[Path]:
    """
    Exporta um arquivo oficial por curso.

    Na versão 4.2, o arquivo-modelo é criado dinamicamente a partir
    das datas informadas. Não existe mais associação de uma coluna
    única a "Segunda", "Quarta" etc.; cada data recebe sua própria
    coluna, o que permite repetir dias da semana sem sobrescrita.
    """
    datas_normalizadas = normalizar_datas_provas(
        datas_provas
    )

    if not datas_normalizadas:
        raise ValueError(
            "O layout oficial exige datas_provas com datas reais."
        )

    if calendario.empty:
        return []

    colunas_obrigatorias = {
        "curso",
        "periodo",
        "cod_turma",
        "data_prova",
        "hora",
        "codigo",
        "materia",
        "turma",
        "sala",
        "professores",
    }
    faltantes = (
        colunas_obrigatorias
        - set(calendario.columns)
    )

    if faltantes:
        raise ValueError(
            "Colunas ausentes para o layout oficial: "
            + ", ".join(
                sorted(faltantes)
            )
        )

    datas_permitidas = {
        formatar_data_br(data_real)
        for data_real
        in datas_normalizadas
    }
    datas_usadas = {
        limpar_texto(valor)
        for valor
        in calendario["data_prova"]
        if limpar_texto(valor)
    }
    datas_fora_modelo = (
        datas_usadas
        - datas_permitidas
    )

    if datas_fora_modelo:
        raise ValueError(
            "O calendário contém datas sem coluna no modelo oficial: "
            + ", ".join(
                sorted(datas_fora_modelo)
            )
        )

    pasta = Path(
        pasta_saida
    )
    pasta.mkdir(
        parents=True,
        exist_ok=True,
    )

    arquivos = []

    for curso in sorted(
        calendario["curso"]
        .dropna()
        .astype(str)
        .unique(),
        key=str.casefold,
    ):
        dados_curso = calendario[
            calendario["curso"]
            .astype(str)
            .eq(curso)
        ].copy()

        wb, ws, linha_atual = (
            criar_arquivo_modelo_oficial(
                curso=curso,
                nome_avaliacao=nome_avaliacao,
                periodo_letivo=periodo_letivo,
                datas_provas=datas_normalizadas,
            )
        )

        for (
            periodo,
            cod_turma,
        ), grupo in dados_curso.groupby(
            [
                "periodo",
                "cod_turma",
            ],
            sort=True,
            dropna=False,
        ):
            linha_atual = (
                preencher_bloco_modelo_oficial(
                    ws=ws,
                    linha_inicial=linha_atual,
                    grupo=grupo,
                    periodo=int(periodo),
                    cod_turma=limpar_texto(
                        cod_turma
                    ),
                    datas_provas=(
                        datas_normalizadas
                    ),
                )
            )

        ultima_coluna = get_column_letter(
            1 + len(datas_normalizadas)
        )
        ws.print_area = (
            f"A1:{ultima_coluna}"
            f"{max(1, linha_atual - 1)}"
        )
        ws.oddFooter.center.text = (
            "Página &P de &N"
        )

        nome_arquivo = (
            f"{nome_seguro_arquivo(nome_avaliacao)}_"
            f"{nome_seguro_arquivo(curso)}_"
            f"{nome_seguro_arquivo(periodo_letivo)}.xlsx"
        )
        caminho = pasta / nome_arquivo
        wb.save(
            caminho
        )
        wb.close()
        arquivos.append(
            caminho
        )

    return arquivos



# ============================================================
# EXCEL — LAYOUT OFICIAL IBMEC (MODELO INSTITUCIONAL)
# ============================================================

# Logo extraída do próprio documento-modelo do IBMEC. Mantê-la
# incorporada torna o script autossuficiente.
LOGO_IBMEC_BASE64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAJoAAAA/CAYAAAAYLtFKAAAAIGNIUk0AAHomAACAhAAA+gAAAIDo"
    "AAB1MAAA6mAAADqYAAAXcJy6UTwAAAAGYktHRAD/AP8A/6C9p5MAAAAJcEhZcwAACxIAAAsSAdLd"
    "fvwAAAAHdElNRQfqCBMVMSt5276dAAAR6UlEQVR42u2ceZBlVXnAf+fc7a39et+me2BmgBkywIiI"
    "qFBBZQlEMLhgRZO4lFETLU0qpso/kkolf6QqS5nESGmSUktMXColKosJRAMaomwCggwDQ88+vcz0"
    "dPfrt971nPxxH72919OvxyHdo/dX1dUz955zvvt997tn+c53GhISEhISEhISEhISEhISEhISEhLW"
    "QpxhPZlLkSlk6ErZpBHgBbjFCsWaS0WB2mjFEjYX63Y0U2AO9jB83R5uvO0q3n7xKJdIgdw/yYv3"
    "Pc7d9z/JfcdPcTRUhButXMLmYV2OJgRiWz8XfOwW/uC338j7cmlyS+/XPGrfeoRvfvrb/NVLE7yg"
    "NXqjFUzYHBjrKZxL0fGeN/LeD/0aHy1k6Vh53zKxzutjWxDiP3WAJ7wAb6MVTNgcyPUUvnCYnW/Y"
    "xTVdOTpXK1PI0vGaC7lq91Yu22jlEjYP63K0oS4GR/vYula5wS6GRnrXLpfwy8O6HE1IDCnXriMF"
    "0mijXMIvD+tyhpkyp07MMbVWuVMlpqfaKJfwy8O6HG1sgv1PvMRj5TqV1cpUXarPHOLpfcf42UYr"
    "l7B5WNeqs+ZRrXnUBjsZ3NLDqGVgCRGHSJRC1zxq//s8/3Pnf/PFvcd4bqOVS9g8rMvRAKaKTByc"
    "4qBtYmdT5PwQv1KncqLI1IPP8v3P/Qf/+OPneViTxNASFjnTLShyKXI7hrhoxxAXSoE8fIKDLx5n"
    "X9mltNFKJWw+ztTRRJt1Nc09Wzt1W9VLOIc5A0cTBoYziJEaWLNo5M+g3Am0CuKqZidmeghhpE9b"
    "T4c1Im8SFcxvtIESzg7mumtIs0DnBR+j86JPrlm2dORfmdn7p2g3DnWk+66lZ/dfYHdcfNp63tyT"
    "zO77S2pT391oAyWcHdbvaAAIAyEthFjeI2oVoSIXXs7cUB7oJUOgjtDKRas6aIEwHKThNDcvLRBn"
    "PH9M2HycoaOtQuhN4878mKAyBoBX/Ak6qi3cD6qHqBz7d4zUAMJIk+q+ilT3lU0Om/ALx9l1tMgd"
    "p3T4TmqT97S8H5T3UizvBUDavXTt/BRO12sSR/vF5+w62iuHQEgTjYUQBqDQKgIC0NErK09KtNII"
    "fLQKOf1qWIC0ECIe+rVSCII26rWDBGEhpBn/GxW3qwPObkazCdKM5WiB1tHZ0GGTO5owEGYe0+7H"
    "zIxgpLcgzSxaBUTeKaL6OKF3AuUX0VGVJoNLG8PqQVqF04qJwiLKPQlIDDuPsHoxMyOY6RGkkUVF"
    "NaL6OEHtKMo/iQrKy2QJaSHMAobTj5kZxUwNIwwHFVaJ6hME9eMobxoVlGA9mcfCQJo5hNWJ6fRh"
    "prZgOD0gHbSqE7knCWvjRP40kV8EVeeMnEHaSLOAtAqYqX6M1DCG3QXCRIUVIneKsD6JCuZQQam1"
    "rU/P5nU0aWSxOy4hO3QLuZF3YOV3IY0UQgi01qAioqCIV3yayvi3qZ98kKB2tGHsGCt9Hp0XfZL8"
    "yDtPK6t05E7m9n8aMz1EdvAWsltuw8rvXCJPoSKXoLyP8tFvUBn/DlH9CFqFCDOPXbiU3PBtZIdv"
    "wcpuQ0h7ST2PoHqQ6vh3qE58G29+H6ja6ZUXAmHksXIXkOl7M9nht2AXLsOwOuKPb8EGIZE3TX3m"
    "ESrj36J+6mGUN7UQTlrbyDbS7sYp7CE7eDPp/jdhZbcjjTRCxrtGWqvGh30Sd/ZJalMPUJ/5MWHt"
    "KDos0abDbU5HE0aGzOCNFC74OOme18fD15LVqxACDBPT6MUcvIF0369Sn36Y4thnqU8/hA7LL5eM"
    "hwDDBiRC2kjDarZCZpTc6Lvo2Po7OF2XNz+PkBhmBqPrCqz8LpzCHub2/x2RP01u+DYKO34fp/Ar"
    "jRejV9RLYxR24+R3kem7lrkX/5bqie8t+yBWKI+RGiI3fBsd2z+EU7gEIeLkB60itPLRWse9qGFh"
    "ZobJZ95BZuAGqhP3Uhz7LF7xp6BPn90sjCx24VV0bHs/uS1vw3R6Fu5ppVChi260IYSNmRomPzJK"
    "bvhW/NKLlI5+hcqxbxLWD7czfVn3XifCSJPuvYZ079VNk/jQnaQ+/UOCyottOVO69xpSPa9vakda"
    "nTiFy7AyW+Phxz1B6E4SedOoqN4UXhHSxMpuw+7YReTN4FcOxHMXacZDmDdNWD2IMFKYqb5mR0tv"
    "Id17DWZmKyooEtUnidwpVFgBIRHSWZAlpY2Z3Y7p9JHqvpLCjo9iZUZRQZnQnSKqT8ZDOTTqyQWn"
    "M9MjGM4AQWWMsH60pZOZmW10XfAJunb+EXb2vIXeSwVl/PJ+3LknCSovoVUdIVMLzyYNByt/MWZm"
    "lLB2kLA2wWq9jTDz5IZvpnv3n5EbvhXDzMYOpjUqrBJUD+DO/IjayYfw5p4idCdBK4SZjm2Y7ifV"
    "8wYQAnfmUVDuWq97c/Zo0rBQoYtfOUDtxPepTf0nQXUMrRVWfif5kXeSHfx1jFT/orMJgdP5Kgrb"
    "P0LknqQ+/SDKO0H58JcpH/4yZmYb3Rf/CU6hOVhs2J1EQQVv7nEq4/dQn/4BkXcq7i1G30Nu5HbM"
    "VO9ieStHbvT2xsuJCKoHqE09QGXyPoLKAaSZJzv8VjrOfx92btviC5aSVM/ryA7fijf/LDosLn39"
    "GKlhCts/TMe2D2LYi2cyIr9I5fg3mR+7A7+8F7TGzG6nsO1D5M9/H1Z6YMFumf7riLxpwvoUYXWs"
    "hXHTZAdvonPnp3A6X73sI1fBPNXJ71Ic+xxe8Seg/AU/sTt20bH9I+RHfxMz1YuUKaRVQBgOeu1p"
    "5+Z0NBX51Gcfpbj/76lN/RfoxS8mrB7AL/6MyJulsOPDGFZ+8VUJQarnteRG3oZf+hmRd6IteVFQ"
    "iQ28/9N4c0/yck8Q1g4S+fNIu4f86O3LXoqQBiry8eaepPjSZ6hM3LNsOAzdCaSZo/OiP0TKRTsb"
    "Vhancw92xy682UcX2zMyZIduJb/13cucTKmQ+vRDFMfuICg9u8QOY8wf+gKG00vHee9FGLEMaTpk"
    "+q/DKz7N/IF/WjGESlJdryZ//geanSzyqE09wNwLf4O/RE5DGn7pOebH7kCaOfJbf6ux+m+bzZlu"
    "HVQPUz785XgLSjd3y5F7jOrE3bizTzRrZKRIdb+OVM/r25bnzT9L6dCX8OaeYOVwE9YO4ZeeR7eY"
    "hwSVAxQPfJ7K8bua5lzKO4FffgHlF5vqmekRrMz2JVdEY+HzFqzMyApdT8bTkdLzTe2EtYPxxNyf"
    "XtH+MOneq7Gy5y+3jd1NZvBmMr3XNk1XgsoBKhP34JdWT1gNKi9RP/kgkTuxjrcZi15vhf8XdFgi"
    "8mdOO8kMSs/hzTzWiKctx8pfhN15ebyV1Y68oIwKiqs8ixvP1Vo8S+THIZbVQhY6qqHD5mxkaXUg"
    "7S4WkhqEQ6r7SlLdr23Ws3oQf35vaxk6IqwdIaweXnZZSImd34Vd2LP0Kk5hD+m+NyKtzPJmlMIr"
    "/hR37lFOHx5R+JWX8Er72rLrUpXXW2HTEAVz+NWDRP5c0z3DymFnt2Okhl7RZ5BmBmnm1l0vXjE6"
    "C35mpAawC5fGMbKVevozRP4swsi2/NHKJ2qR5WKmt2DnLwDRGLalhd2xG6ewu6msimoElTGi2rG1"
    "7e6eIKwdi3t43XbMbnPO0dpFeaeI3KllE/WXMRqBx7B29AxabhNhgbTPpOKy/1mZEazs9pZbcabd"
    "S7rvWpzO1udkzcwohtOsvzDzjT3lDDosIa0urNx2DLuz2Y7+DKE73lb8TQdF/NLzeMWnY9uqtgLQ"
    "57ijhWVU2DqjV1qFOLp9DiDtbgynv+W9dN/VpPuuXn+bhoW0CkgzSxSWMOweDKd1DmEUzBN5s221"
    "GwVF5g/+C6UjX0ErD6J6O9XObUfTOoiVbYGQqTUTLDcLwsggrWzLe1FQjT8mvf79TB1WF4bOWEbr"
    "YV5H7rIsmzVaBVVDr7W7sZxz29FArzpPEEIubKNsdsQq6e1aKdxTD1M++lXUwm5H+wS1cSJvZsEe"
    "q2dUv+Jp8+e6oxmwijNpHaAjf53tbQxah43siGbC+jjVE99DtRkTXFWGCladgwlpLi4aXhnO3VUn"
    "gDBSCNl6yFFhDRVUN/oR20IFJVTQvHoWUiLMLNLInkGrK+2xeghHmgUMp/uVVPHcdjTD6lzVQMqf"
    "I/Km19nixhD5Jwnr4611tHvaOgi0toxZQneqZdzRsHsxUyO8giPcuetoQlgYqSFMZ7DpnlYRoTux"
    "6svbbIS1cfzyS6gWQ72ZjkMf6zuxZrLSaXRQIqwdigPhK5BWHju/E3PFrsRZZJM6mpCINeYMMr0F"
    "u+Pipig3xEFOv7wfFcxwLqCCWfziMy1jfmZmK07X5QirzVCNTMeB2c7LEcuCySF+6QW8+b3LUpkg"
    "3re1O/c0tu3a6NVkupEY6axdtlFjI+y69lPZPVjZHRjWkm2a5QVIdb+WdIv9zHg75Vm8ucc5Zw4h"
    "6wi3+CT16R829WqGlW2kZV2zRnA4TpbM9L+Z3j1/Te+r/oF033UsdRx/fm+cHNliW8zO7yQ7dCtW"
    "fiesumHekDFwAx07Pk6692qQqXY03JyrTiuzlY5tv4vWEdWJu4m8U6DdOHffzOAU9pAfeVfL86GR"
    "P0t9+ge4s09ttBrrIqweoDJ5H07Xq3EKexBysRNwOi+n4/z3o/wZvPln4vjYwkckQaYwU72k+66n"
    "sOP3cLquQPmzOJ2XNxJB46C2Cmapn/we6a4rSQ9cvywJVBo2mcGbiPxZ5g98nqh6CLU0NVw4GE4P"
    "mYGb6Lzwozgdl1KZvC8ePeafWUu9s+toQtgYTh9mZmtDsSoqLC5uSMs0htWJMCyk3Y208q2Pb2qN"
    "nb+Qnkv+nMzAjdQm78Mvv4DWCqfzMnJbbifVc1VTnEyFNWpT91MZ/87a6dKbDR1RP/kgpfQWui76"
    "Y8zM6IJ+0rDJDt2C6QxQOnon7sxjqLAKCKSZw+7YTW7Lb5DufzOG3UUUFKmM30X52NcWnKwhBHf2"
    "cUpH7sRID2F37F7mbKbTRce292PndlA+8m9488+gIhchLazsNnJb3k526BaMVB+RP0dYPbwQp1uD"
    "s+toZuZ8Oi/4BPmt7wagNvUA84e+iA7i7Y1U9xV0nPcBzOx5CGE1fjcP36F7AhXOY6ZHyA7dRG74"
    "5tO/I6VQUZna1PeZP/TPBOXnORfRYYny0a8DMs7czZ6PNOOhSRoW6b43kOp5HSpoZLcIE8PpXsjJ"
    "UyokdKeoHL+L4tgdhJX9zUKUT23iXkDQeeEnsAuXIs3swj6rYWbJDt1EZuAGIr+ICopIM4u0e+KE"
    "VBUS1qcoH/sG82N3ELnH21HtDB1NqzjIqJd3R8JMYzdy5wHC6hGEYccnwgDD6iXVfSXWkiFPtQhU"
    "ujM/ojJxL5mB68n0vQnD6YlTlpcMJ1prtI7QYZ3Im6I6dT+lg19o5FO1nptpHbWUF+eanT49RusQ"
    "pVbMXdQaGQxaN+qFKy6vfnRNBbOUDn2JsHY0PjPQdTmG1bWgv5ASw+nEcDoX7KAiHxVW4pyyY9+g"
    "fOxrRO7qf3FTqRrVibsIakcobPsg6b5rMVJ9CJlBSBMhBEIamKkeSPXEtlYBoTdLUBmjfOSrlI99"
    "HeW3HT5av6NpHRLUjuDNPLpmWb86tiw6H/mzeMWfrho4XKhX2kd9+iFqU/eTGbiR3Ja3xqegzPzi"
    "uc7IJ/Kmceeeojp5L+7sI0sOpbR47sglrB7Em3mk6V5Q3tdyggyA0ETuBN7MY035bX7lQMtwwaK+"
    "p/CKTxO5kyuuFwnrE6v6tlZVqpP34M79hEz/9WSGbsLOX4y0C40zozJ2cB2iohph7Rj1Uw9Tnbhn"
    "9dy1JhkB3uwjTJeew+l6DdnBm3C6rsBMDSHNbGNBIECrWEb9OPXpH1KduBt//vm2ZCy14noKL9YR"
    "NqKN9Bitg/jvbyxMKE2EdFhrtat12KjX2EiWaYzUwIIRFs51upOocL7NQ8QSpB2/qFbytL96O9JG"
    "tFrKawXaj/VsZSlpgbCb9RUarfwlOflrmdzBSPVjZc/DcAZiG+q4h4nqxwnrE+iows+3ypYII4uR"
    "6sdMDcaJmcJEh1XCehyT1OH8zykjISEhISEhISEhISEhISEhISEhISEhISEhIeFs8X9iVr9y2IyN"
    "2wAAAABJRU5ErkJggg=="
)

COR_IBMEC_AZUL = "17365D"
COR_IBMEC_AMARELO = "FFC000"
COR_CINZA_TEXTO = "7A7A7A"
COR_CINZA_BORDA = "7F7F7F"
COR_BRANCA = "FFFFFF"
COR_PRETA = "000000"
COR_CONFLITO = "FFC7CE"

LARGURA_COLUNA_HORARIO = 18.0
LARGURA_TOTAL_DIAS_MODELO = 5 * 26.0
# Linha em que a paginação "n/total" é escrita nas abas que cabem
# numa folha A4 (a grade empilhada usa a primeira linha livre).
LINHA_RODAPE_MODELO = 55

FAIXAS_HORARIOS_MODELO = {
    "07:30": "7:30 às 9:20",
    "09:50": "9:50 às 11:40",
    "13:30": "13:30 às 15:20",
    "15:50": "15:50 às 17:40",
    "18:40": "18:40 às 20:30",
    "20:40": "20:40 às 22:30",
}


def cor_rgb(
    valor: str,
) -> str:
    return (
        valor
        if len(valor) == 8
        else f"FF{valor}"
    )


def fonte_modelo(
    *,
    tamanho: float = 8,
    negrito: bool = False,
    italico: bool = False,
    cor: str = COR_PRETA,
) -> Font:
    return Font(
        name="Arial",
        size=tamanho,
        bold=negrito,
        italic=italico,
        color=cor_rgb(cor),
    )


def fonte_inline_modelo(
    *,
    tamanho: float = 7.5,
    negrito: bool = False,
    cor: str = COR_PRETA,
) -> InlineFont:
    """Versão inline de fonte_modelo, para trechos de rich text."""
    return InlineFont(
        rFont="Arial",
        sz=tamanho,
        b=negrito,
        color=Color(
            rgb=cor_rgb(cor)
        ),
    )


def lado_borda_modelo(
    estilo: str | None,
    cor: str = COR_PRETA,
) -> Side:
    return Side(
        style=estilo,
        color=(
            cor_rgb(cor)
            if estilo
            else None
        ),
    )


def borda_modelo(
    *,
    esquerda: str | None = None,
    direita: str | None = None,
    superior: str | None = None,
    inferior: str | None = None,
) -> Border:
    return Border(
        left=lado_borda_modelo(
            esquerda
        ),
        right=lado_borda_modelo(
            direita
        ),
        top=lado_borda_modelo(
            superior
        ),
        bottom=lado_borda_modelo(
            inferior
        ),
    )


def rotulo_curso_oficial(
    curso: str,
    nome_avaliacao: str,
) -> str:
    aliases = {
        "cdia": (
            "CIÊNCIA DE DADOS E "
            "INTELIGÊNCIA ARTIFICIAL"
        ),
        "ec": "ENGENHARIA DE COMPUTAÇÃO",
        "ecomp": "ENGENHARIA DE COMPUTAÇÃO",
        "engenharia de computacao": (
            "ENGENHARIA DE COMPUTAÇÃO"
        ),
        "ep": "ENGENHARIA DE PRODUÇÃO",
        "eprod": "ENGENHARIA DE PRODUÇÃO",
        "engenharia de producao": (
            "ENGENHARIA DE PRODUÇÃO"
        ),
        "es": "ENGENHARIA DE SOFTWARE",
        "esw": "ENGENHARIA DE SOFTWARE",
        "engenharia de software": (
            "ENGENHARIA DE SOFTWARE"
        ),
    }

    curso_limpo = limpar_texto(
        curso
    )
    chave = normalizar_nome_comparacao(
        curso_limpo
    )
    nome = aliases.get(
        chave,
        curso_limpo.upper(),
    )

    if (
        "FARIA LIMA"
        not in normalizar_nome_comparacao(
            nome
        ).upper()
    ):
        nome = (
            f"{nome} - FARIA LIMA"
        )

    avaliacao = limpar_texto(
        nome_avaliacao
    ).upper()

    return (
        f"{avaliacao} - {nome}"
        if avaliacao
        else nome
    )


def rotulo_data_modelo(
    data_real: date,
) -> str:
    dia = nome_dia(
        data_real
    ).upper()

    if dia in {
        "SEGUNDA",
        "TERÇA",
        "QUARTA",
        "QUINTA",
        "SEXTA",
    }:
        dia = f"{dia}-FEIRA"

    return (
        f"{dia}\n"
        f"{data_real.strftime('%d/%m')}"
    )


def adicionar_logo_modelo(
    ws,
) -> None:
    logo = XLImage(
        BytesIO(
            b64decode(
                LOGO_IBMEC_BASE64
            )
        )
    )
    logo.width = 154
    logo.height = 63
    logo.anchor = "H1"
    ws.add_image(
        logo
    )


def configurar_pagina_modelo(
    ws,
    datas_provas: list[date],
) -> None:
    quantidade_datas = len(
        datas_provas
    )
    largura_dia = (
        LARGURA_TOTAL_DIAS_MODELO
        / quantidade_datas
    )

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = (
        ws.PAPERSIZE_A4
    )
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = (
        True
    )

    # Mesmas margens registradas no arquivo-modelo original.
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    ws.column_dimensions[
        "A"
    ].width = LARGURA_COLUNA_HORARIO

    for indice in range(
        2,
        quantidade_datas + 2,
    ):
        ws.column_dimensions[
            get_column_letter(indice)
        ].width = largura_dia

    # Duas colunas auxiliares mantêm o logo e a paginação na mesma
    # posição relativa do modelo original.
    primeira_auxiliar = (
        quantidade_datas + 2
    )
    for indice in (
        primeira_auxiliar,
        primeira_auxiliar + 1,
    ):
        ws.column_dimensions[
            get_column_letter(indice)
        ].width = 11.0

    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 28

    ultima_auxiliar = get_column_letter(
        primeira_auxiliar + 1
    )
    ws.print_area = (
        f"A1:{ultima_auxiliar}{LINHA_RODAPE_MODELO}"
    )


def criar_cabecalho_modelo(
    ws,
    *,
    curso: str,
    nome_avaliacao: str,
    periodo_letivo: str,
    datas_provas: list[date],
) -> None:
    quantidade_datas = len(
        datas_provas
    )
    ultima_data = (
        quantidade_datas + 1
    )

    # A distribuição A:E / F:G é a extensão proporcional das
    # mesclagens A:D / E:F do modelo original de cinco dias.
    divisao = max(
        2,
        ultima_data - 2,
    )

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=divisao,
    )
    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=divisao,
    )
    ws.merge_cells(
        start_row=1,
        start_column=divisao + 1,
        end_row=1,
        end_column=ultima_data,
    )
    ws.merge_cells(
        start_row=2,
        start_column=divisao + 1,
        end_row=2,
        end_column=ultima_data,
    )

    fill_azul = PatternFill(
        "solid",
        fgColor=cor_rgb(
            COR_IBMEC_AZUL
        ),
    )
    fill_branco = PatternFill(
        "solid",
        fgColor=cor_rgb(
            COR_BRANCA
        ),
    )
    borda_fina = borda_modelo(
        esquerda="thin",
        direita="thin",
        superior="thin",
        inferior="thin",
    )

    for linha in (1, 2):
        for coluna in range(
            1,
            ultima_data + 1,
        ):
            celula = ws.cell(
                row=linha,
                column=coluna,
            )
            celula.border = borda_fina
            celula.fill = (
                fill_azul
                if linha == 1
                else fill_branco
            )

    ws.cell(
        row=1,
        column=divisao + 1,
        value="ANO E PERÍODO",
    )
    ws.cell(
        row=2,
        column=1,
        value=rotulo_curso_oficial(
            curso,
            nome_avaliacao,
        ),
    )
    ws.cell(
        row=2,
        column=divisao + 1,
        value=periodo_letivo,
    )

    ws.cell(
        row=1,
        column=divisao + 1,
    ).font = fonte_modelo(
        negrito=True,
        cor=COR_BRANCA,
    )

    for referencia in (
        (2, 1),
        (2, divisao + 1),
    ):
        celula = ws.cell(
            row=referencia[0],
            column=referencia[1],
        )
        celula.font = fonte_modelo(
            negrito=True,
        )

    for referencia in (
        (1, divisao + 1),
        (2, 1),
        (2, divisao + 1),
    ):
        ws.cell(
            row=referencia[0],
            column=referencia[1],
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # Cabeçalho temporal no mesmo azul do modelo.
    for coluna in range(
        1,
        ultima_data + 1,
    ):
        celula = ws.cell(
            row=4,
            column=coluna,
        )
        celula.fill = fill_azul
        celula.font = fonte_modelo(
            negrito=True,
            cor=COR_BRANCA,
        )
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        celula.border = borda_modelo(
            esquerda=(
                "thin"
                if coluna == 1
                else None
            ),
            direita=(
                "thin"
                if coluna == ultima_data
                else None
            ),
            superior="thin",
            inferior="thin",
        )

    for coluna, data_real in enumerate(
        datas_provas,
        start=2,
    ):
        ws.cell(
            row=4,
            column=coluna,
            value=rotulo_data_modelo(
                data_real
            ),
        )

    adicionar_logo_modelo(
        ws
    )


def titulo_bloco_modelo(
    periodo: int,
    cod_turma: str,
) -> str:
    titulo = (
        f"{periodo}º Período"
    )

    if cod_turma:
        titulo += (
            f" | Turma {cod_turma}"
        )

    return titulo


def adicionar_linha_periodo_modelo(
    ws,
    linha: int,
    ultima_coluna: int,
    periodo: int,
    cod_turma: str,
) -> None:
    ws.merge_cells(
        start_row=linha,
        start_column=1,
        end_row=linha,
        end_column=ultima_coluna,
    )

    celula = ws.cell(
        row=linha,
        column=1,
        value=titulo_bloco_modelo(
            periodo,
            cod_turma,
        ),
    )
    celula.fill = PatternFill(
        "solid",
        fgColor=cor_rgb(
            COR_IBMEC_AMARELO
        ),
    )
    celula.font = fonte_modelo(
        negrito=True
    )
    celula.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    celula.border = borda_modelo(
        esquerda="thin",
        direita="thin",
        superior="thin",
        inferior="thin",
    )

    for coluna in range(
        2,
        ultima_coluna + 1,
    ):
        ws.cell(
            row=linha,
            column=coluna,
        ).fill = PatternFill(
            "solid",
            fgColor=cor_rgb(
                COR_IBMEC_AMARELO
            ),
        )

    ws.row_dimensions[
        linha
    ].height = 18


def texto_prova_modelo(
    prova: pd.Series,
) -> list[TextBlock]:
    materia = materia_sem_codigo(
        prova["materia"],
        prova["codigo"],
    ).upper()
    codigo = limpar_texto(
        prova["codigo"]
    ).upper()
    turma = limpar_texto(
        prova["turma"]
    )
    sala = (
        limpar_texto(
            prova["sala"]
        )
        or SALA_INDEFINIDA
    )
    professores = limpar_texto(
        prova["professores"]
    ).upper()

    # Matéria, sala e turma saem em negrito; o restante segue a
    # fonte normal da célula.
    return [
        TextBlock(
            fonte_inline_modelo(
                negrito=True
            ),
            f"{materia}\n\n",
        ),
        TextBlock(
            fonte_inline_modelo(),
            f"{codigo}\n",
        ),
        TextBlock(
            fonte_inline_modelo(
                negrito=True
            ),
            f"Sala {sala}       Turma {turma}\n",
        ),
        TextBlock(
            fonte_inline_modelo(),
            professores,
        ),
    ]


def adicionar_linha_horario_modelo(
    ws,
    *,
    linha: int,
    horario: str,
    grupo: pd.DataFrame,
    datas_provas: list[date],
) -> None:
    ultima_coluna = (
        1 + len(datas_provas)
    )

    celula_hora = ws.cell(
        row=linha,
        column=1,
        value=FAIXAS_HORARIOS_MODELO.get(
            horario,
            horario,
        ),
    )
    celula_hora.fill = PatternFill(
        "solid",
        fgColor=cor_rgb(
            COR_IBMEC_AMARELO
        ),
    )
    celula_hora.font = fonte_modelo(
        negrito=True
    )
    celula_hora.alignment = Alignment(
        horizontal="left",
        vertical="top",
    )
    celula_hora.border = borda_modelo(
        esquerda="thin",
        direita="thin",
        superior="thin",
        inferior="thin",
    )

    maior_quantidade = 1

    for coluna, data_real in enumerate(
        datas_provas,
        start=2,
    ):
        chave_data = formatar_data_br(
            data_real
        )
        provas = grupo[
            (
                grupo["data_prova"]
                == chave_data
            )
            & (
                grupo["hora"]
                == horario
            )
        ]

        celula = ws.cell(
            row=linha,
            column=coluna,
        )
        celula.fill = PatternFill(
            "solid",
            fgColor=cor_rgb(
                (
                    COR_CONFLITO
                    if len(provas) > 1
                    else COR_BRANCA
                )
            ),
        )
        celula.font = fonte_modelo(
            tamanho=7.5,
        )
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
            shrink_to_fit=True,
        )
        celula.border = borda_modelo(
            esquerda=(
                "thin"
                if coluna == 2
                else "dotted"
            ),
            direita=(
                "thin"
                if coluna == ultima_coluna
                else "dotted"
            ),
            superior="dotted",
            inferior="dotted",
        )

        if not provas.empty:
            trechos: list[TextBlock] = []

            for _, prova in provas.iterrows():
                if trechos:
                    trechos.append(
                        TextBlock(
                            fonte_inline_modelo(),
                            "\n\n",
                        )
                    )

                trechos.extend(
                    texto_prova_modelo(
                        prova
                    )
                )

            celula.value = CellRichText(
                *trechos
            )

        maior_quantidade = max(
            maior_quantidade,
            len(provas),
        )

    ws.row_dimensions[
        linha
    ].height = (
        58 * maior_quantidade
    )


def horarios_do_grupo_modelo(
    grupo: pd.DataFrame,
) -> list[str]:
    encontrados = {
        limpar_texto(valor)
        for valor
        in grupo["hora"]
        if limpar_texto(valor)
    }

    ordem = {
        horario: indice
        for indice, horario
        in enumerate(
            ORDEM_HORARIOS
        )
    }

    return sorted(
        encontrados,
        key=lambda horario: (
            ordem.get(
                horario,
                len(ordem),
            ),
            horario,
        ),
    )


def blocos_curso_modelo(
    dados_curso: pd.DataFrame,
) -> list[dict[str, Any]]:
    blocos = []

    for (
        periodo,
        cod_turma,
    ), grupo in dados_curso.groupby(
        [
            "periodo",
            "cod_turma",
        ],
        sort=True,
        dropna=False,
    ):
        horarios = horarios_do_grupo_modelo(
            grupo
        )

        if not horarios:
            continue

        blocos.append(
            {
                "periodo": int(periodo),
                "cod_turma": limpar_texto(
                    cod_turma
                ),
                "grupo": grupo.copy(),
                "horarios": horarios,
                "linhas": 1 + len(horarios),
            }
        )

    return blocos


def paginar_blocos_modelo(
    blocos: list[dict[str, Any]],
) -> list[list[dict[str, Any]]]:
    """Mantém todos os blocos do curso empilhados numa aba só.

    O modelo antigo quebrava a grade em várias abas de no máximo
    MAX_LINHAS_GRADE_POR_PAGINA linhas; hoje a grade inteira fica
    numa única página do Excel e a quebra passa a ser apenas de
    impressão (fitToWidth com as linhas 1:4 repetidas no topo).
    """
    if not blocos:
        return []

    return [list(blocos)]


def ajustar_impressao_grade_modelo(
    ws,
    *,
    datas_provas: list[date],
    ultima_linha: int,
) -> None:
    """Estica a área de impressão até o fim da grade empilhada.

    Como a grade inteira mora numa aba só, ela quase sempre passa da
    altura de uma folha A4: a impressão deixa de ser forçada em uma
    página vertical e o cabeçalho (linhas 1 a 4) passa a se repetir.
    """
    ws.linha_rodape_modelo = max(
        LINHA_RODAPE_MODELO,
        ultima_linha + 1,
    )

    ultima_auxiliar = get_column_letter(
        len(datas_provas) + 3
    )
    ws.print_area = (
        f"A1:{ultima_auxiliar}"
        f"{ws.linha_rodape_modelo}"
    )

    if ws.linha_rodape_modelo > LINHA_RODAPE_MODELO:
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = "1:4"


def criar_pagina_grade_modelo(
    wb: Workbook,
    *,
    numero: int,
    curso: str,
    nome_avaliacao: str,
    periodo_letivo: str,
    datas_provas: list[date],
    blocos: list[dict[str, Any]],
) -> Any:
    ws = wb.create_sheet(
        f"Página {numero}"
    )
    configurar_pagina_modelo(
        ws,
        datas_provas,
    )
    criar_cabecalho_modelo(
        ws,
        curso=curso,
        nome_avaliacao=nome_avaliacao,
        periodo_letivo=periodo_letivo,
        datas_provas=datas_provas,
    )

    ultima_coluna = (
        1 + len(datas_provas)
    )
    linha = 5

    for bloco in blocos:
        adicionar_linha_periodo_modelo(
            ws,
            linha,
            ultima_coluna,
            bloco["periodo"],
            bloco["cod_turma"],
        )
        linha += 1

        for horario in bloco[
            "horarios"
        ]:
            adicionar_linha_horario_modelo(
                ws,
                linha=linha,
                horario=horario,
                grupo=bloco["grupo"],
                datas_provas=datas_provas,
            )
            linha += 1

    ajustar_impressao_grade_modelo(
        ws,
        datas_provas=datas_provas,
        ultima_linha=linha - 1,
    )

    return ws


def criar_pagina_observacoes_modelo(
    wb: Workbook,
    *,
    numero: int,
    curso: str,
    nome_avaliacao: str,
    periodo_letivo: str,
    datas_provas: list[date],
) -> Any:
    ws = wb.create_sheet(
        f"Página {numero}"
    )
    configurar_pagina_modelo(
        ws,
        datas_provas,
    )
    criar_cabecalho_modelo(
        ws,
        curso=curso,
        nome_avaliacao=nome_avaliacao,
        periodo_letivo=periodo_letivo,
        datas_provas=datas_provas,
    )

    quantidade_datas = len(
        datas_provas
    )
    ultima_data = (
        quantidade_datas + 1
    )
    primeira_auxiliar = (
        ultima_data + 1
    )
    ultima_auxiliar = (
        primeira_auxiliar + 1
    )

    coluna_aviso_final = max(
        3,
        ultima_data - 3,
    )
    coluna_coord_inicio = (
        coluna_aviso_final + 1
    )

    ws.merge_cells(
        start_row=6,
        start_column=1,
        end_row=9,
        end_column=coluna_aviso_final,
    )
    ws.merge_cells(
        start_row=6,
        start_column=coluna_coord_inicio,
        end_row=7,
        end_column=ultima_data,
    )
    ws.merge_cells(
        start_row=8,
        start_column=coluna_coord_inicio,
        end_row=9,
        end_column=ultima_data,
    )
    ws.merge_cells(
        start_row=6,
        start_column=primeira_auxiliar,
        end_row=9,
        end_column=ultima_auxiliar,
    )

    aviso = (
        "A Grade de disciplinas apresentada não se constitui em um "
        "compromisso formal das Coordenações Acadêmicas, podendo "
        "sofrer alterações durante o período letivo em relação à "
        "disponibilização de turmas, seus horários e alocação do "
        "corpo docente."
    )

    celula_aviso = ws.cell(
        row=6,
        column=1,
        value=aviso,
    )
    celula_aviso.font = fonte_modelo(
        italico=True,
        cor=COR_CINZA_TEXTO,
    )
    celula_aviso.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    celula_coord = ws.cell(
        row=6,
        column=coluna_coord_inicio,
        value="COORDENAÇÃO ACADÊMICA",
    )
    celula_coord.font = fonte_modelo(
        negrito=True,
        italico=True,
        cor=COR_CINZA_TEXTO,
    )
    celula_coord.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    celula_curso = ws.cell(
        row=8,
        column=coluna_coord_inicio,
        value=rotulo_curso_oficial(
            curso,
            "",
        ),
    )
    celula_curso.font = fonte_modelo(
        negrito=True,
        italico=True,
        cor=COR_CINZA_TEXTO,
    )
    celula_curso.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    celula_atualizacao = ws.cell(
        row=6,
        column=primeira_auxiliar,
        value=(
            "Atualizado em:\n"
            f"{date.today().strftime('%d/%m/%Y')}"
        ),
    )
    celula_atualizacao.font = fonte_modelo(
        cor=COR_CINZA_TEXTO,
    )
    celula_atualizacao.alignment = Alignment(
        horizontal="right",
        vertical="center",
        wrap_text=True,
    )

    for linha in range(
        6,
        10,
    ):
        ws.row_dimensions[
            linha
        ].height = 17

    return ws


def numerar_paginas_modelo(
    wb: Workbook,
    quantidade_datas: int,
) -> None:
    total = len(
        wb.worksheets
    )
    coluna = get_column_letter(
        quantidade_datas + 3
    )

    for numero, ws in enumerate(
        wb.worksheets,
        start=1,
    ):
        celula = ws[
            f"{coluna}"
            f"{getattr(ws, 'linha_rodape_modelo', LINHA_RODAPE_MODELO)}"
        ]
        celula.value = (
            f"{numero}/{total}"
        )
        celula.font = fonte_modelo(
            tamanho=9,
            negrito=True,
        )
        celula.alignment = Alignment(
            horizontal="right",
            vertical="bottom",
        )


def exportar_calendarios_oficiais(
    calendario: pd.DataFrame,
    nome_avaliacao: str,
    periodo_letivo: str,
    datas_provas: Iterable[str | date | datetime] | None,
    pasta_saida: str | Path,
) -> list[Path]:
    """
    Exporta um arquivo por curso no mesmo layout institucional IBMEC.

    A única extensão estrutural em relação ao modelo original é que
    B:G passa a conter uma coluna para cada uma das seis datas da AS.
    A chave é a data completa, por isso quartas e quintas repetidas
    permanecem independentes.

    As colunas seguem a ordem cronológica, mas giradas para começar
    na segunda-feira, mesmo quando a primeira data do calendário cai
    numa quinta.
    """
    datas_normalizadas = normalizar_datas_provas(
        datas_provas
    )

    if datas_normalizadas:
        datas_normalizadas = (
            ordenar_datas_por_dia_semana(
                datas_normalizadas
            )
        )

    if not datas_normalizadas:
        raise ValueError(
            "O layout oficial exige datas_provas com datas reais."
        )

    if calendario.empty:
        return []

    colunas_obrigatorias = {
        "curso",
        "periodo",
        "cod_turma",
        "data_prova",
        "hora",
        "codigo",
        "materia",
        "turma",
        "sala",
        "professores",
    }
    faltantes = (
        colunas_obrigatorias
        - set(calendario.columns)
    )

    if faltantes:
        raise ValueError(
            "Colunas ausentes para o layout oficial: "
            + ", ".join(
                sorted(faltantes)
            )
        )

    datas_permitidas = {
        formatar_data_br(
            data_real
        )
        for data_real
        in datas_normalizadas
    }
    datas_usadas = {
        limpar_texto(valor)
        for valor
        in calendario["data_prova"]
        if limpar_texto(valor)
    }
    datas_fora_modelo = (
        datas_usadas
        - datas_permitidas
    )

    if datas_fora_modelo:
        raise ValueError(
            "O calendário contém datas sem coluna no modelo oficial: "
            + ", ".join(
                sorted(
                    datas_fora_modelo
                )
            )
        )

    pasta = Path(
        pasta_saida
    )
    pasta.mkdir(
        parents=True,
        exist_ok=True,
    )

    arquivos = []

    for curso in sorted(
        calendario["curso"]
        .dropna()
        .astype(str)
        .unique(),
        key=str.casefold,
    ):
        dados_curso = calendario[
            calendario["curso"]
            .astype(str)
            .eq(curso)
        ].copy()

        blocos = blocos_curso_modelo(
            dados_curso
        )
        paginas = paginar_blocos_modelo(
            blocos
        )

        if not paginas:
            continue

        wb = Workbook()
        del wb[
            wb.sheetnames[0]
        ]

        for numero, blocos_pagina in enumerate(
            paginas,
            start=1,
        ):
            criar_pagina_grade_modelo(
                wb,
                numero=numero,
                curso=curso,
                nome_avaliacao=nome_avaliacao,
                periodo_letivo=periodo_letivo,
                datas_provas=datas_normalizadas,
                blocos=blocos_pagina,
            )

        criar_pagina_observacoes_modelo(
            wb,
            numero=len(paginas) + 1,
            curso=curso,
            nome_avaliacao=nome_avaliacao,
            periodo_letivo=periodo_letivo,
            datas_provas=datas_normalizadas,
        )
        numerar_paginas_modelo(
            wb,
            len(datas_normalizadas),
        )

        nome_arquivo = (
            f"{nome_seguro_arquivo(nome_avaliacao)}_"
            f"{nome_seguro_arquivo(curso)}_"
            f"{nome_seguro_arquivo(periodo_letivo)}.xlsx"
        )
        caminho = pasta / nome_arquivo
        wb.save(
            caminho
        )
        wb.close()
        arquivos.append(
            caminho
        )

    return arquivos

# ============================================================
# EXPORTAÇÃO
# ============================================================

def salvar_resultados(
    caminho_saida: str | Path,
    calendario: pd.DataFrame,
    conflitos: pd.DataFrame,
    diagnostico: pd.DataFrame,
    concentracao_professores: pd.DataFrame,
    ofertas_df: pd.DataFrame,
    nao_alocadas: pd.DataFrame,
    excluidas: pd.DataFrame,
    validacao_geral: pd.DataFrame,
    validacao_ofertas: pd.DataFrame,
    parametros: pd.DataFrame,
    datas_provas: list[date] | None,
    nome_avaliacao: str | None,
) -> None:
    with pd.ExcelWriter(
        caminho_saida,
        engine="openpyxl",
    ) as writer:

        calendario.to_excel(
            writer,
            sheet_name="Dados do Calendário",
            index=False,
        )

        conflitos.to_excel(
            writer,
            sheet_name="Conflitos",
            index=False,
        )

        diagnostico.to_excel(
            writer,
            sheet_name="Diagnóstico",
            index=False,
        )

        concentracao_professores.to_excel(
            writer,
            sheet_name="Concentração Professores",
            index=False,
        )

        ofertas_df.to_excel(
            writer,
            sheet_name="Ofertas",
            index=False,
        )

        nao_alocadas.to_excel(
            writer,
            sheet_name="Não Alocadas",
            index=False,
        )

        excluidas.to_excel(
            writer,
            sheet_name="Excluídas da Alocação",
            index=False,
        )

        validacao_geral.to_excel(
            writer,
            sheet_name="Validação",
            index=False,
        )

        validacao_ofertas.to_excel(
            writer,
            sheet_name="Validação Ofertas",
            index=False,
        )

        parametros.to_excel(
            writer,
            sheet_name="Parâmetros",
            index=False,
        )

        criar_aba_calendario_visual(
            writer,
            calendario,
            datas_provas,
            nome_avaliacao,
            nao_alocadas,
            excluidas,
        )

        for nome_aba in [
            "Dados do Calendário",
            "Conflitos",
            "Diagnóstico",
            "Concentração Professores",
            "Ofertas",
            "Não Alocadas",
            "Excluídas da Alocação",
            "Validação",
            "Validação Ofertas",
            "Parâmetros",
        ]:
            formatar_aba_tabela(
                writer.book[
                    nome_aba
                ]
            )


# ============================================================
# FUNÇÃO PRINCIPAL
# ============================================================

def gerar_calendario_provas(
    arquivo_grade: str | Path,
    arquivo_saida: str | Path = "calendario_provas.xlsx",
    datas_provas: Iterable[str | date | datetime] | None = None,
    nome_avaliacao: str | None = None,
    exportar_oficial: bool = False,
    periodo_letivo: str | None = None,
    pasta_oficial: str | Path = "OFICIAL",
    concentrar_por_professor: bool = True,
    excluir_ed_eletivas: bool = False,
    dias_substitutos: dict[str, tuple[str, ...]] | None = None,
    hora_dia_substituto: str | None = HORA_DIA_SUBSTITUTO,
    hora_noturna_padrao: str | None = HORA_NOTURNA_PADRAO,
    unir_turmas_conjuntas: bool = True,
) -> dict[str, Any]:
    """
    Fluxo:

    1. Lê o CSV/XLSX combinado dos cursos.
    2. Valida os dados.
    3. Opcionalmente separa (ED) e (E) (excluir_ed_eletivas=True).
       Por padrão elas entram na alocação como qualquer outra
       disciplina.
    4. Consolida ofertas por codigo + turma + agenda + professor.
       Nomes diferentes da disciplina não rompem o compartilhamento;
       professores diferentes sempre geram ofertas distintas.
    5. Cada oferta guarda todos os grupos de alunos que atende:
       (curso, período, cod_turma).
    6. O solver escolhe uma única prova por oferta.
    7. A mesma decisão afeta simultaneamente todos os grupos
       atendidos pela oferta.
    8. Minimiza duas ou mais provas no mesmo dia para cada
       curso + período + turma. Turmas diferentes do mesmo período
       não disputam entre si.
    9. Sem piorar o objetivo dos alunos, concentra as provas de cada
       professor no menor número possível de dias e aproxima os dias
       restantes.
    10. Opcionalmente, escreve um .xlsx por curso no layout oficial
       IBMEC (exportar_oficial=True), pronto para distribuição.
    """
    print(
        "1. Carregando grade..."
    )

    df = carregar_grade(
        arquivo_grade
    )

    print(
        f"   {len(df)} linha(s) carregada(s)."
    )

    print(
        "2. Validando dados gerais..."
    )

    validacao_geral = (
        validar_grade(
            df
        )
    )

    print(
        "3. Tratando Estudo Dirigido e Eletivas..."
        if excluir_ed_eletivas
        else "3. Estudo Dirigido e Eletivas entram na alocação."
    )

    (
        df_alocacao,
        df_excluidas,
    ) = (
        separar_disciplinas_excluidas(
            df,
            excluir=excluir_ed_eletivas,
        )
    )

    excluidas = (
        criar_dataframe_excluidas(
            df_excluidas
        )
    )

    if excluir_ed_eletivas:
        print(
            f"   {len(excluidas)} oferta(s) "
            "(ED)/(E) excluída(s)."
        )

    print(
        "4. Validando ofertas por codigo + turma + agenda + professor..."
    )

    validacao_ofertas = (
        validar_consistencia_ofertas(
            df_alocacao
        )
    )

    garantir_consistencia_critica(
        validacao_ofertas
    )

    print(
        "5. Consolidando ofertas acadêmicas..."
    )

    ofertas = construir_ofertas(
        df_alocacao
    )

    compartilhadas = [
        oferta
        for oferta in ofertas
        if len(
            oferta.grupos
        ) > 1
    ]

    print(
        f"   {len(ofertas)} oferta(s) única(s) identificada(s)."
    )

    print(
        f"   {len(compartilhadas)} oferta(s) compartilhada(s) "
        "entre mais de um curso/período/turma."
    )

    grupos_alunos = {
        grupo
        for oferta in ofertas
        for grupo in oferta.grupos
    }

    print(
        f"   {len(grupos_alunos)} grupo(s) de alunos "
        "(curso + período + turma)."
    )

    datas_normalizadas = (
        normalizar_datas_provas(
            datas_provas
        )
    )

    if datas_normalizadas:
        print(
            "6. Datas permitidas:"
        )

        for data_real in (
            datas_normalizadas
        ):
            print(
                f"   {formatar_data_br(data_real)} "
                f"({nome_dia(data_real)})"
            )
    else:
        print(
            "6. Sem datas reais: usando dias da semana."
        )

    print(
        "7. Executando otimização global..."
    )

    substituicoes = {
        dia_aula: dias_alvo
        for dia_aula, dias_alvo
        in (
            dias_substitutos
            or {}
        ).items()
        if datas_normalizadas
        and dia_aula not in {
            nome_dia(data_real)
            for data_real
            in datas_normalizadas
        }
    }

    for dia_aula, dias_alvo in sorted(
        substituicoes.items(),
        key=lambda item: (
            ORDEM_DIAS[item[0]]
        ),
    ):
        print(
            f"   Sem data de {dia_aula}: essas provas "
            "podem cair em "
            + ", ".join(dias_alvo)
            + (
                f" (as noturnas às {hora_dia_substituto} "
                f"em vez de {hora_noturna_padrao}, "
                "as diurnas no horário da aula)."
                if hora_dia_substituto
                else " (mesmo horário)."
            )
        )

    if unir_turmas_conjuntas:
        for membros in agrupar_turmas_conjuntas(
            ofertas
        ):
            print(
                "   Turma conjunta (mesma aula, uma prova só): "
                + " = ".join(
                    f"{oferta.codigo}"
                    for oferta in membros
                )
                + f" — sala {sorted(membros[0].salas)[0]}."
            )

    for membros in detectar_choques_de_professor(
        ofertas
    ):
        print(
            "   ATENÇÃO: "
            + " / ".join(
                sorted(membros[0].professores)
            )
            + " tem "
            + " e ".join(
                f"{oferta.codigo} (sala "
                f"{sorted(oferta.salas)[0]})"
                for oferta in membros
            )
            + " no mesmo horário, em salas diferentes. "
            "Conferir a grade."
        )

    (
        solver,
        escolhas,
        opcoes_por_oferta,
        metricas_otimizacao,
    ) = otimizar_calendario(
        ofertas,
        datas_provas=(
            datas_normalizadas
        ),
        concentrar_por_professor=(
            concentrar_por_professor
        ),
        dias_substitutos=(
            dias_substitutos
        ),
        hora_dia_substituto=(
            hora_dia_substituto
        ),
        hora_noturna_padrao=(
            hora_noturna_padrao
        ),
        unir_turmas_conjuntas=(
            unir_turmas_conjuntas
        ),
    )

    nao_alocadas = (
        criar_dataframe_nao_alocadas(
            ofertas,
            opcoes_por_oferta,
            datas_normalizadas,
        )
    )

    if not nao_alocadas.empty:
        print(
            f"   {len(nao_alocadas)} oferta(s) sem "
            "data elegível ficarão para tratamento manual."
        )

    print(
        "8. Construindo resultados..."
    )

    calendario = (
        criar_dataframe_calendario(
            ofertas,
            escolhas,
        )
    )

    conflitos = (
        criar_dataframe_conflitos(
            calendario
        )
    )

    diagnostico = (
        criar_dataframe_diagnostico(
            calendario
        )
    )

    concentracao_professores = (
        criar_dataframe_concentracao_professores(
            ofertas,
            escolhas,
        )
    )

    ofertas_df = (
        criar_dataframe_ofertas(
            ofertas,
            escolhas,
            opcoes_por_oferta,
        )
    )

    parametros = (
        criar_dataframe_parametros(
            datas_normalizadas,
            nome_avaliacao,
            concentrar_por_professor,
        )
    )

    print(
        "9. Salvando Excel..."
    )

    salvar_resultados(
        arquivo_saida,
        calendario,
        conflitos,
        diagnostico,
        concentracao_professores,
        ofertas_df,
        nao_alocadas,
        excluidas,
        validacao_geral,
        validacao_ofertas,
        parametros,
        datas_normalizadas,
        nome_avaliacao,
    )

    arquivos_oficiais = []

    if exportar_oficial:
        print(
            "10. Exportando layout oficial por curso..."
        )

        arquivos_oficiais = (
            exportar_calendarios_oficiais(
                calendario=calendario,
                nome_avaliacao=(
                    nome_avaliacao
                    or "Calendario"
                ),
                periodo_letivo=(
                    periodo_letivo
                    or periodo_letivo_padrao(
                        datas_normalizadas
                    )
                ),
                datas_provas=datas_normalizadas,
                pasta_saida=pasta_oficial,
            )
        )

        for caminho in arquivos_oficiais:
            print(
                f"   {caminho}"
            )

    print()
    print(
        "Calendário gerado com sucesso!"
    )
    print(
        f"Arquivo: {arquivo_saida}"
    )
    print(
        "Objetivo dos alunos (mesmo critério da v3): "
        f"{metricas_otimizacao.objetivo_alunos}"
    )

    if concentrar_por_professor:
        print(
            "Professores considerados na concentração "
            "(com 2 ou mais provas): "
            f"{metricas_otimizacao.professores_considerados}"
        )
        print(
            "Dias adicionais usados pelos professores: "
            f"{metricas_otimizacao.dias_adicionais_professores}"
        )
        print(
            "Intervalo total das provas dos professores: "
            f"{metricas_otimizacao.intervalo_total_professores} dia(s)"
        )

        if not (
            metricas_otimizacao
            .concentracao_por_professor_otimizada
        ):
            print(
                "ATENÇÃO: o limite de tempo foi atingido antes da "
                "etapa de concentração; foi preservada a solução "
                "dos alunos encontrada na primeira etapa."
            )
    else:
        print(
            "Concentração por professor: desativada."
        )

    if conflitos.empty:
        print(
            "Nenhum conflito encontrado entre "
            "as ofertas alocadas."
        )
    else:
        print(
            f"Foram encontrados {len(conflitos)} "
            "dia(s) com conflito."
        )

    if not nao_alocadas.empty:
        print(
            f"ATENÇÃO: {len(nao_alocadas)} oferta(s) "
            "não foram alocadas automaticamente."
        )

    if not excluidas.empty:
        print(
            f"INFO: {len(excluidas)} oferta(s) (ED)/(E) "
            "foram excluídas do solver."
        )

    return {
        "calendario": calendario,
        "conflitos": conflitos,
        "diagnostico": diagnostico,
        "concentracao_professores": (
            concentracao_professores
        ),
        "ofertas": ofertas_df,
        "nao_alocadas": nao_alocadas,
        "excluidas": excluidas,
        "validacao": validacao_geral,
        "validacao_ofertas": validacao_ofertas,
        "parametros": parametros,
        "metricas_otimizacao": (
            metricas_otimizacao
        ),
        "arquivos_oficiais": arquivos_oficiais,
    }


# ============================================================
# MAIN — EDITE SOMENTE ESTA PARTE
# ============================================================

if __name__ == "__main__":
    # --------------------------------------------------------
    # 1. ARQUIVOS
    # --------------------------------------------------------

    # CSV gerado no passo 1 do README (a partir do seu PDF/planilha).
    # Troque por "exemplo/grade_exemplo.csv" para testar a instalação.
    ARQUIVO_GRADE = (
        "grade.csv"
    )

    # Planilha de trabalho: calendário, conflitos, diagnóstico.
    ARQUIVO_SAIDA = (
        "AP1_MINHA_UNIDADE.xlsx"
    )

    # Pasta onde sai um .xlsx por curso, no layout institucional.
    PASTA_OFICIAL = (
        "OFICIAL/MINHA_UNIDADE"
    )

    EXPORTAR_OFICIAL = True

    # --------------------------------------------------------
    # 2. IDENTIFICAÇÃO
    # --------------------------------------------------------

    # AP1, AP2 ou AS.
    NOME_AVALIACAO = (
        "AP1"
    )

    PERIODO_LETIVO = (
        "2026.2"
    )

    # --------------------------------------------------------
    # 3. DATAS DAS PROVAS
    #
    # Datas reais, no formato AAAA-MM-DD. Cada data vira uma coluna
    # do calendário oficial; duas datas do mesmo dia da semana são
    # colunas independentes.
    #
    # As colunas mantêm a ordem cronológica, mas giradas para
    # começar na segunda-feira: um calendário que começa na quinta
    # é apresentado a partir da segunda, com a quinta e a sexta no
    # fim da tabela.
    #
    # Use None para o modo antigo, por dias da semana, sem datas.
    # --------------------------------------------------------

    # Para testar com exemplo/grade_exemplo.csv, use:
    # DATAS_PROVAS = [
    #     "2026-05-11",  # Segunda
    #     "2026-05-12",  # Terça
    #     "2026-05-13",  # Quarta
    #     "2026-05-14",  # Quinta
    #     "2026-05-15",  # Sexta
    # ]

    DATAS_PROVAS = [
        "2026-09-24",  # Quinta
        "2026-09-25",  # Sexta
        "2026-09-28",  # Segunda
        "2026-09-29",  # Terça
        "2026-09-30",  # Quarta
    ]

    # --------------------------------------------------------
    # 4. REGRAS DE NEGÓCIO
    #
    # Os padrões abaixo refletem a prática da unidade Faria Lima.
    # Veja a seção 8 do README antes de mudar.
    # --------------------------------------------------------

    # Estudo Dirigido (ED) e Eletivas (E) entram na alocação junto
    # com as demais disciplinas. Use True para deixá-las de fora do
    # solver, na aba "Excluídas da Alocação".
    EXCLUIR_ED_ELETIVAS = False

    # Dias de aula sem nenhuma data correspondente em DATAS_PROVAS.
    # Só vale quando NENHUM dia de aula da disciplina tem data: uma
    # disciplina de quarta e sexta faz a prova na quarta. Use {}
    # para desligar.
    DIAS_SUBSTITUTOS = DIAS_SUBSTITUTOS_PADRAO

    # Mantém as regras dos alunos como prioridade absoluta e usa a
    # concentração por professor apenas para desempatar calendários
    # igualmente bons para os estudantes.
    CONCENTRAR_POR_PROFESSOR = True

    # Une automaticamente disciplinas de códigos diferentes que são
    # a mesma aula (mesmo professor + mesma agenda + mesma sala).
    UNIR_TURMAS_CONJUNTAS = True

    resultado = (
        gerar_calendario_provas(
            arquivo_grade=(
                ARQUIVO_GRADE
            ),
            arquivo_saida=(
                ARQUIVO_SAIDA
            ),
            datas_provas=(
                DATAS_PROVAS
            ),
            nome_avaliacao=(
                NOME_AVALIACAO
            ),
            exportar_oficial=(
                EXPORTAR_OFICIAL
            ),
            periodo_letivo=(
                PERIODO_LETIVO
            ),
            pasta_oficial=(
                PASTA_OFICIAL
            ),
            concentrar_por_professor=(
                CONCENTRAR_POR_PROFESSOR
            ),
            excluir_ed_eletivas=(
                EXCLUIR_ED_ELETIVAS
            ),
            dias_substitutos=(
                DIAS_SUBSTITUTOS
            ),
            unir_turmas_conjuntas=(
                UNIR_TURMAS_CONJUNTAS
            ),
        )
    )

    print()
    print(
        "CONFLITOS:"
    )
    print(
        resultado[
            "conflitos"
        ]
    )
